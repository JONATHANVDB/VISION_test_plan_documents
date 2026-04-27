"""
Surgically refresh the four Chapter "Product Specifications & Acceptance
Criteria" tables in Vision_Series_Production_Test_Plan.md from
chapter7_config.xlsx, leaving everything else (sections, paragraphs, the
pixel-array defect table, references, etc.) untouched.

For each product the script:
  1. Reads the configured rows from chapter7_config.xlsx (Include = TRUE).
  2. For IC products (CXP/IMG/PSU): overrides LSL / USL with the latest
     values from the corresponding check_limits.xlsx (matched by Mxxxx)
     and converts them from base SI to the display unit, exactly the same
     way `chapter7_config_manager.py report` does.
  3. Locates the existing markdown table by its title marker
     `*Table: <Product> Key Specifications*`, finds the contiguous block
     of `|`-rows that follows it, and replaces only that block with the
     freshly built one.

Run:
    python update_chapter7_tables.py
"""

import os
import re
import sys

from openpyxl import load_workbook

# Reuse helpers from the existing manager so the rendering stays consistent
from chapter7_config_manager import (
    CXP_CHECK_LIMITS,
    IMG_CHECK_LIMITS,
    PSU_CHECK_LIMITS,
    read_check_limits,
    _read_ic_sheet,
    _read_vis_sheet,
    _fmt,
    _convert_lsl_usl,
    display_unit,
)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
CONFIG_XLSX = os.path.join(BASE_DIR, "chapter7_config.xlsx")
MD_FILE = os.path.join(BASE_DIR, "Vision_Series_Production_Test_Plan.md")

TABLE_TITLES = {
    "IMG002X1": "*Table: MAG-IMG002x1-NC Key Specifications*",
    "CXP00002": "*Table: MAG-CXP00002-NP Key Specifications*",
    "PSU00001": "*Table: MAG-PSU00001-NP Key Specifications*",
    "VIS100xx": "*Table: MAG-VIS100xx-N Key Specifications*",
}

# Mxxxx whose check_limits LSL/USL are boolean/error-count indicators that
# should be rendered as "pass/fail" rather than scaled by the display unit.
# (e.g. mode_success = 1 is not "1 MHz", and pixel_errors USL=0 is "no errors")
PASS_FAIL_OVERRIDES = {
    ("IMG002X1", "M0130"): ("pass/fail", "-"),  # T0024 Frame modes (mode_success)
    ("CXP00002", "M0156"): ("pass/fail", "-"),  # T0023 Internal Test Patterns (pixel_errors)
}


def build_ic_table_block(rows, product_key):
    """Build the markdown table body for an IC product (no title line)."""
    lines = [
        "| Parameter | Datasheet Label | DS Min | DS Typ | DS Max | Unit | LSL | USL | Comment |",
        "| :--- | :--- | :--- | :--- | :--- | :--- | :--- | :--- | :--- |",
    ]
    prev_section = None
    for r in rows:
        section = r.get("Section", "")
        if section and section != prev_section:
            lines.append(f"| **{section}** | | | | | | | | |")
            prev_section = section

        param = r.get("Parameter", "-")
        ds_label = _fmt(r.get("Datasheet_Label", ""))
        ds_min = _fmt(r.get("DS_Min", ""))
        ds_typ = _fmt(r.get("DS_Typ", ""))
        ds_max = _fmt(r.get("DS_Max", ""))
        raw_unit = r.get("Unit", "")
        unit_for_scale = _fmt(raw_unit)
        mxxxx = (r.get("Mxxxx", "") or "").strip()
        override = PASS_FAIL_OVERRIDES.get((product_key, mxxxx))
        if override:
            lsl, usl = override
        else:
            lsl = _convert_lsl_usl(r.get("LSL", ""), unit_for_scale)
            usl = _convert_lsl_usl(r.get("USL", ""), unit_for_scale)
        unit_display = _fmt(display_unit(raw_unit))
        comment = _fmt(r.get("Comment", ""))

        lines.append(
            f"| {param} | {ds_label} | {ds_min} | {ds_typ} | {ds_max} | {unit_display} | {lsl} | {usl} | {comment} |"
        )
    return "\n".join(lines)


def build_vis_table_block(rows):
    """Build the markdown table body for the VIS module (no LSL/USL columns)."""
    lines = [
        "| Parameter | Datasheet Label | DS Min | DS Typ | DS Max | Unit | Comment |",
        "| :--- | :--- | :--- | :--- | :--- | :--- | :--- |",
    ]
    prev_section = None
    for r in rows:
        section = r.get("Section", "")
        if section and section != prev_section:
            lines.append(f"| **{section}** | | | | | | |")
            prev_section = section

        param = r.get("Parameter", "-")
        ds_label = _fmt(r.get("Datasheet_Label", ""))
        ds_min = _fmt(r.get("DS_Min", ""))
        ds_typ = _fmt(r.get("DS_Typ", ""))
        ds_max = _fmt(r.get("DS_Max", ""))
        unit_display = _fmt(display_unit(r.get("Unit", "")))
        comment = _fmt(r.get("Comment", ""))

        lines.append(
            f"| {param} | {ds_label} | {ds_min} | {ds_typ} | {ds_max} | {unit_display} | {comment} |"
        )
    return "\n".join(lines)


def replace_table_block(content: str, title_marker: str, new_block: str) -> str:
    """Replace the markdown table that follows `title_marker` with `new_block`.

    The existing table is identified as the contiguous run of lines starting
    with `|` that begins at the first such line after the marker (skipping
    any blank lines).  Everything else in the file is untouched.
    """
    lines = content.split("\n")
    title_idx = None
    for i, line in enumerate(lines):
        if line.strip() == title_marker:
            title_idx = i
            break
    if title_idx is None:
        print(f"  [WARN] Marker not found: {title_marker}")
        return content

    # Find first table line after the marker
    j = title_idx + 1
    while j < len(lines) and lines[j].strip() == "":
        j += 1
    if j >= len(lines) or not lines[j].lstrip().startswith("|"):
        print(f"  [WARN] No table found after marker: {title_marker}")
        return content
    table_start = j

    # Find end of contiguous `|` block
    k = table_start
    while k < len(lines) and lines[k].lstrip().startswith("|"):
        k += 1
    table_end = k  # exclusive

    new_lines = lines[:table_start] + new_block.split("\n") + lines[table_end:]
    return "\n".join(new_lines)


def main():
    if not os.path.exists(CONFIG_XLSX):
        print(f"Config not found: {CONFIG_XLSX}")
        sys.exit(1)
    if not os.path.exists(MD_FILE):
        print(f"Markdown not found: {MD_FILE}")
        sys.exit(1)

    print("Re-reading check_limits for fresh LSL/USL …")
    fresh_limits = {
        "IMG002X1": read_check_limits(IMG_CHECK_LIMITS),
        "CXP00002": read_check_limits(CXP_CHECK_LIMITS),
        "PSU00001": read_check_limits(PSU_CHECK_LIMITS),
    }
    for k, v in fresh_limits.items():
        print(f"  {k}: {len(v)} Mxxxx entries")

    print("Loading config …")
    wb = load_workbook(CONFIG_XLSX, data_only=True)

    tables = {}
    for sn in ["IMG002X1", "CXP00002", "PSU00001"]:
        if sn not in wb.sheetnames:
            print(f"  [WARN] Sheet missing: {sn}")
            continue
        rows = _read_ic_sheet(wb[sn])
        limits = fresh_limits.get(sn, {})
        for r in rows:
            mxxxx = r.get("Mxxxx", "").strip()
            if mxxxx and mxxxx in limits:
                lsl = limits[mxxxx].get("lsl")
                usl = limits[mxxxx].get("usl")
                r["LSL"] = str(lsl) if lsl is not None else ""
                r["USL"] = str(usl) if usl is not None else ""
            else:
                r["LSL"] = ""
                r["USL"] = ""
        tables[sn] = rows

    if "VIS100xx" in wb.sheetnames:
        tables["VIS100xx"] = _read_vis_sheet(wb["VIS100xx"])
    wb.close()

    print("Building markdown table blocks …")
    blocks = {}
    for sn, rows in tables.items():
        if sn == "VIS100xx":
            blocks[sn] = build_vis_table_block(rows)
        else:
            blocks[sn] = build_ic_table_block(rows, sn)

    print(f"Patching {os.path.basename(MD_FILE)} …")
    with open(MD_FILE, "r", encoding="utf-8") as f:
        content = f.read()

    for sn, block in blocks.items():
        title = TABLE_TITLES[sn]
        before = content
        content = replace_table_block(content, title, block)
        status = "updated" if content != before else "unchanged"
        print(f"  {sn:9} ({title}) -> {status}")

    with open(MD_FILE, "w", encoding="utf-8") as f:
        f.write(content)

    print("Done.")


if __name__ == "__main__":
    main()
