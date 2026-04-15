"""
Chapter 7 Configuration Manager
================================
Generates and reads back an Excel configuration file that serves as the
single source of truth for Chapter 7 (Product Specifications & Acceptance
Criteria) of the Vision Series Test Plan.

Usage
-----
  python chapter7_config_manager.py generate   # create / overwrite config xlsx
  python chapter7_config_manager.py report      # read config → update markdown → build docx
"""

import os
import re
import sys
from copy import copy

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ---------------------------------------------------------------------------
# Paths (all relative to this script)
# ---------------------------------------------------------------------------
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
CONFIG_XLSX = os.path.join(BASE_DIR, "chapter7_config.xlsx")
MD_FILE = os.path.join(BASE_DIR, "Vision_Series_Test_Plan.md")

PROJECTS_DIR = os.path.join(BASE_DIR, os.pardir)
CXP_CHECK_LIMITS = os.path.join(
    PROJECTS_DIR, "proxima-product-test-sw", "sw", "f_post_processing", "check_limits.xlsx"
)
PSU_CHECK_LIMITS = os.path.join(
    PROJECTS_DIR, "DCDC-tester-python-sw", "dcdc_tester_python_sw",
    "f_post_processing", "check_limits.xlsx"
)
IMG_CHECK_LIMITS = os.path.join(
    PROJECTS_DIR, "Vision-MEGACAM-python-sw", "vision_megacam_python_sw",
    "f_post_processing", "check_limits.xlsx"
)
IMG_EO_PARAMS_XLSX = os.path.join(
    BASE_DIR, "test plans", "IMG002xx_electr_optical_parameters.xlsx"
)

# ---------------------------------------------------------------------------
# Styling constants
# ---------------------------------------------------------------------------
HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL_GENERAL = PatternFill("solid", fgColor="2F5496")
HEADER_FILL_DS = PatternFill("solid", fgColor="4472C4")       # blue  → datasheet
HEADER_FILL_CL = PatternFill("solid", fgColor="ED7D31")       # orange → check_limits
SECTION_FILL = PatternFill("solid", fgColor="D9E2F3")
SECTION_FONT = Font(name="Calibri", bold=True, size=11)
DATA_FONT = Font(name="Calibri", size=11)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

# Column layout for IC products (CXP, IMG, PSU)
IC_COLUMNS = [
    ("Include",         12, HEADER_FILL_GENERAL),
    ("Section",         20, HEADER_FILL_GENERAL),
    ("Parameter",       40, HEADER_FILL_GENERAL),
    ("Datasheet_Label", 22, HEADER_FILL_GENERAL),
    ("DS_Min",          14, HEADER_FILL_DS),
    ("DS_Typ",          14, HEADER_FILL_DS),
    ("DS_Max",          14, HEADER_FILL_DS),
    ("Unit",            10, HEADER_FILL_GENERAL),
    ("Mxxxx",           12, HEADER_FILL_GENERAL),
    ("Test_Proc",       12, HEADER_FILL_GENERAL),
    ("Test_Name",       28, HEADER_FILL_GENERAL),
    ("LSL",             14, HEADER_FILL_CL),
    ("USL",             14, HEADER_FILL_CL),
    ("Comment",         50, HEADER_FILL_GENERAL),
]

# Column layout for VIS module (no Mxxxx / check_limits)
VIS_COLUMNS = [
    ("Include",         12, HEADER_FILL_GENERAL),
    ("Section",         20, HEADER_FILL_GENERAL),
    ("Parameter",       40, HEADER_FILL_GENERAL),
    ("Datasheet_Label", 22, HEADER_FILL_GENERAL),
    ("DS_Min",          14, HEADER_FILL_DS),
    ("DS_Typ",          14, HEADER_FILL_DS),
    ("DS_Max",          14, HEADER_FILL_DS),
    ("Unit",            10, HEADER_FILL_GENERAL),
    ("Test_Proc",       12, HEADER_FILL_GENERAL),
    ("Test_Name",       28, HEADER_FILL_GENERAL),
    ("Comment",         50, HEADER_FILL_GENERAL),
]

# ---------------------------------------------------------------------------
# Helper: read check_limits.xlsx → dict  {Mxxxx: {lsl, usl, parameter}}
# ---------------------------------------------------------------------------
def read_check_limits(xlsx_path):
    """Return {Mxxxx_id: {'lsl': ..., 'usl': ..., 'param': ...}}."""
    limits = {}
    if not os.path.exists(xlsx_path):
        print(f"  [WARN] check_limits not found: {xlsx_path}")
        return limits
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return limits

    header = [str(c).strip() if c else "" for c in rows[0]]
    idx = {name: i for i, name in enumerate(header)}
    pid_col = idx.get("PARAMETER_ID", idx.get("parameter_id", None))
    lsl_col = idx.get("LSL", idx.get("lsl", None))
    usl_col = idx.get("USL", idx.get("usl", None))
    par_col = idx.get("PARAMETER", idx.get("parameter", None))
    if pid_col is None:
        print(f"  [WARN] No PARAMETER_ID column in {xlsx_path}")
        return limits

    mxxxx_re = re.compile(r'M\d{4}')
    for row in rows[1:]:
        pid = str(row[pid_col]) if row[pid_col] else ""
        match = mxxxx_re.search(pid)
        if not match:
            continue
        mid = match.group(0)
        lsl_val = row[lsl_col] if lsl_col is not None else None
        usl_val = row[usl_col] if usl_col is not None else None
        par_val = str(row[par_col]).strip() if par_col is not None and row[par_col] else ""
        if mid not in limits:
            limits[mid] = {"lsl": lsl_val, "usl": usl_val, "param": par_val}
    wb.close()
    return limits

# ---------------------------------------------------------------------------
# Datasheet parameter definitions (seed data)
# Each entry: (section, parameter, ds_label, ds_min, ds_typ, ds_max, unit,
#              mxxxx, test_proc, comment)
# Section rows (headers) have section != "" and parameter == ""
# ---------------------------------------------------------------------------

CXP_PARAMS = [
    # Section headers use parameter="" to mark them as category rows
    ("Recommended operating conditions", "VDD3V3 supply voltage", "VDD3V3", "2.97", "3.3", "3.63", "V", "", "", ""),
    ("Recommended operating conditions", "Operation temperature (junction)", "Tj", "-40", "", "85", "°C", "", "", ""),
    ("Current consumption", "3.3V total supply current (average)", "IVDD3V3_TOTAL_AVG", "", "82.35", "", "mA", "", "T0007", "Sum of analog + digital + PLL + camera"),
    ("Current consumption", "3.3V digital supply current (average)", "IVDD3V3_DIG_AVG", "14.45", "18.6", "28", "mA", "M0041", "T0007", ""),
    ("Current consumption", "3.3V analog supply current (average)", "IVDD3V3_ANA_AVG", "", "49.8", "", "mA", "M0035", "T0007", "Active mode measurement"),
    ("Current consumption", "3.3V PLL supply current (average)", "IVDD3V3_PLL_AVG", "", "13.93", "", "mA", "", "T0007", "Included in total, no separate Mxxxx"),
    ("Power consumption", "Total power consumption", "", "", "271.55", "", "mW", "", "T0007", "Derived from IDD x VDD"),
    ("Logic level inputs (SPI / GPIO)", "VIL (GPIO input low threshold)", "VIL", "390", "470", "635", "mV", "M0016", "T0004", ""),
    ("Logic level inputs (SPI / GPIO)", "VIH (GPIO input high threshold)", "VIH", "443", "550", "715", "mV", "M0017", "T0004", ""),
    ("Logic level inputs (SPI / GPIO)", "Hysteresis (GPIO)", "Hysteresis", "38", "80", "117", "mV", "M0018", "T0004", ""),
    ("Logic level inputs (Video DV<11:0>)", "VIL (Video input low threshold)", "VIL", "1.147", "1.328", "1.689", "V", "M0019", "T0004", ""),
    ("Logic level inputs (Video DV<11:0>)", "VIH (Video input high threshold)", "VIH", "1.482", "1.714", "2.282", "V", "M0020", "T0004", ""),
    ("Logic level inputs (Video DV<11:0>)", "Hysteresis (Video)", "Hysteresis", "309", "387", "593", "mV", "M0021", "T0004", ""),
    ("Digital output levels (GPIO)", "VOL (GPIO output low voltage)", "VOL", "", "", "300", "mV", "M0029", "T0005", ""),
    ("Digital output levels (GPIO)", "VOH (GPIO output high voltage)", "VOH", "VDD-0.3", "", "", "V", "M0030", "T0005", ""),
    ("Digital output levels (GPIO)", "IOL (GPIO output low current)", "IOL", "20", "", "", "mA", "M0031", "T0006", ""),
    ("Digital output levels (GPIO)", "IOH (GPIO output high current)", "IOH", "-20", "", "", "mA", "M0032", "T0006", ""),
    ("Data rates", "Parallel video data rate", "DRin,par", "", "", "100", "MHz", "", "Functional", ""),
    ("Data rates", "Downlink data rate (CXP)", "DRdown", "", "1.25", "", "Gbps", "", "T0022/T0023", ""),
    ("Data rates", "Uplink data rate (CXP)", "DRup", "", "20.833", "", "Mbps", "", "T0020/T0021", ""),
    ("Data rates", "Transceiver load impedance", "ZL", "", "75", "", "Ω", "", "", ""),
    ("Crystal oscillator", "Crystal frequency", "fXTAL", "10", "20", "40", "MHz", "M0131", "T0013", ""),
    ("CAM LDO outputs", "VDD1V8_CAM supply voltage", "V_VDD1V8_CAM", "1.7", "1.8", "1.9", "V", "M0085", "T0009", ""),
    ("CAM LDO outputs", "VDD1V8_CAM load current", "I_VDD1V8_CAM", "", "", "10", "mA", "", "T0012", ""),
    ("CAM LDO outputs", "VDD1V2_CAM reference voltage", "V_VDD1V2_CAM", "1.15", "1.2", "1.25", "V", "M0088", "T0009", ""),
    ("Radiation hardness", "Total Ionizing Dose (TID)", "", "> 1 MGy (100 Mrad)", "", "", "Si", "", "By design + lot acceptance", ""),
    ("Radiation hardness", "Single Event Effects (SEE)", "", "> 60", "", "", "MeV.cm²/mg", "", "By design", ""),
]

IMG_PARAMS = [
    ("Recommended operating conditions", "1.8V digital supply voltage (VDDD)", "VDDD", "1.62", "1.8", "1.98", "V", "", "", ""),
    ("Recommended operating conditions", "3.3V analog supply voltage (VDDA)", "VDDA", "3.135", "3.3", "3.63", "V", "", "", ""),
    ("Recommended operating conditions", "3.3V IO supply voltage (VDDIO)", "VDDIO", "3.135", "3.3", "3.63", "V", "", "", ""),
    ("Recommended operating conditions", "Pixel reset voltage (VD_RST)", "VD_RST", "2.5", "3.3", "3.3", "V", "", "", ""),
    ("Recommended operating conditions", "External reference clock (REFCLK)", "REFCLK", "10", "20", "40", "MHz", "", "", ""),
    ("Recommended operating conditions", "SPI clock speed (S_SCLK)", "S_SCLK", "", "10", "20", "MHz", "", "", ""),
    ("Recommended operating conditions", "Operation temperature (junction, pre-irradiation)", "Tj", "0", "25", "85", "°C", "", "", ""),
    ("Current consumption", "VDDD current (active, FHD25 10-bit)", "I_VDDD", "", "205", "1000", "mA", "M0160", "T0007", "Max 1000 mA only at startup/reset"),
    ("Current consumption", "VDDA+VDDIO+VD_RST current (active, FHD25 10-bit)", "I_VDDA (combined)", "", "124", "", "mA", "", "T0007", "Combined 3.3V supplies; individual: M0161, M0162, M0025"),
    ("Clocking and control signal thresholds", "TRIG_IN input threshold low", "TRIG_in_l", "0.74", "1.07", "1.4", "V", "", "T0004", ""),
    ("Clocking and control signal thresholds", "TRIG_IN input threshold high", "TRIG_in_h", "1.56", "1.89", "2.39", "V", "M0006", "T0004", "trig_in_vih"),
    ("Clocking and control signal thresholds", "TRIG_IN input hysteresis", "TRIG_in_hyst", "0.66", "0.82", "1.15", "V", "", "T0004", ""),
    ("Clocking and control signal thresholds", "RSTB input threshold (no hysteresis)", "RSTB_th_lvl", "", "1.65", "", "V", "M0009", "T0004", "rstb_vih"),
    ("SPI interface", "SPI MISO output voltage (no load)", "SPI_out", "2.97", "3.3", "3.63", "V", "M0020", "T0005", "pin_voh"),
    ("SPI interface", "SPI input threshold low", "SPI_in_l", "0.74", "1.07", "1.4", "V", "M0017", "T0004", "s_mosi_vil (representative)"),
    ("SPI interface", "SPI input threshold high", "SPI_in_h", "1.56", "1.89", "2.39", "V", "M0016", "T0004", "s_mosi_vih (representative)"),
    ("SPI interface", "SPI input hysteresis", "SPI_in_hyst", "0.66", "0.82", "1.15", "V", "M0018", "T0004", "s_mosi_hysteresis (representative)"),
    ("Digital video output interface", "VOH (no resistive load)", "VOH", "2.97", "VDDIO", "3.63", "V", "M0020", "T0005", "pin_voh"),
    ("Digital video output interface", "VOL (no resistive load)", "VOL", "", "0 (gnd)", "", "V", "M0019", "T0005", "pin_vol"),
    ("Digital video output interface", "Pixel output clock speed", "PIXCLK_speed", "", "50", "100", "MHz", "", "Functional", ""),
    ("Digital video output interface", "Data output line speed", "DV_X_speed", "", "", "50", "MHz", "", "Functional", ""),
    ("Analog signals", "Bandgap reference voltage (VDD12BG)", "Vvdd12bg", "1.14", "1.2", "1.26", "V", "M0029", "T0009", "bgr_vref"),
    ("Electro-optical characteristics", "Effective number of pixels", "", "", "1920 (H) x 1080 (V)", "", "", "", "", ""),
    ("Electro-optical characteristics", "Pixel size", "", "", "5um x 5um", "", "", "", "", ""),
    ("Electro-optical characteristics", "Number of black rows", "", "", "64", "", "", "", "", "32 at the top; 32 at the bottom"),
    ("Electro-optical characteristics", "Shutter type", "", "", "Rolling shutter", "", "", "", "", ""),
    ("Electro-optical characteristics", "Full well charge", "", "", "94000", "", "e-", "", "", "Preliminary measurement results pre-irradiation"),
    ("Electro-optical characteristics", "Conversion gain", "", "", "0.01", "", "mV/eV", "", "", "Preliminary measurement results pre-irradiation"),
    ("Electro-optical characteristics", "Responsivity", "", "", "0.009", "", "DN/e-", "", "", "Preliminary measurement results pre-irradiation"),
    ("Electro-optical characteristics", "Temporal noise", "", "", "66", "", "e-", "", "", "Preliminary measurement results pre-irradiation"),
    ("Electro-optical characteristics", "Dynamic range", "", "", "62", "", "dB", "", "", "Preliminary measurement results pre-irradiation"),
    ("Electro-optical characteristics", "SNR_MAX", "", "", "49.5", "", "dB", "", "", "Preliminary measurement results pre-irradiation"),
    ("Electro-optical characteristics", "Dark current (DC)", "", "", "0.05", "", "fA", "", "", "Preliminary measurement results pre-irradiation"),
    ("Electro-optical characteristics", "Dark current non uniformity (DCNU)", "", "", "42.8", "", "e-/s", "", "", "Preliminary measurement results pre-irradiation"),
    ("Electro-optical characteristics", "Dark signal non uniformity (DSNU)", "", "", "587.2", "", "e-", "", "", "Preliminary measurement results pre-irradiation"),
    ("Electro-optical characteristics", "Photo response non uniformity (PRNU)", "", "", "< tba", "", "% RMS", "", "", "To be added once characterized"),
    ("Electro-optical characteristics", "Color filters", "", "", "BayerRG (Bayer RGGB)", "", "", "", "", "In the case of the color version"),
    ("Electro-optical characteristics", "Programmable features", "", "", "Sensor parameters", "", "", "", "", "Exposure time, sub-sampling, X-Y mirroring, ADC resolution"),
    ("Electro-optical characteristics", "ADC resolution", "", "", "10 and 12 bit", "", "", "", "", "Selectable"),
    ("Electro-optical characteristics", "Interface", "", "", "10/12b parallel output", "", "", "", "", "CMOS push-pull drivers"),
    ("Electro-optical characteristics", "Cover glass lid", "Corning 7980 0F", "", "", "", "", "", "", "Fused silica with double-side AR coating. Ravg < 1.75% @ 300-800 nm, AOI 0deg"),
    ("Radiation hardness", "Total Ionizing Dose (TID)", "", "> 1 MGy", "", "", "Si", "", "By design + lot acceptance", ""),
    ("Radiation hardness", "Single Event Effects (SEE)", "", "> 62.5", "", "", "MeV.cm²/mg", "", "By design", ""),
]

PSU_PARAMS = [
    # Recommended operating conditions (Table 1)
    ("Recommended operating conditions", "Input voltage (PVin, Vin)", "PVin, Vin", "5", "", "11", "V", "", "", ""),
    ("Recommended operating conditions", "Output voltage (Vout)", "Vout", "0.9", "", "5", "V", "", "", ""),
    ("Recommended operating conditions", "Conversion ratio", "Vout/Vin", "2", "", "10", "", "", "", ""),
    ("Recommended operating conditions", "Output current (active cooling)", "Iout", "0", "", "4", "A", "", "", ""),
    ("Recommended operating conditions", "Output power (active cooling)", "Pout", "0", "", "10", "W", "", "", ""),
    ("Recommended operating conditions", "Switching frequency", "fsw", "1.5", "", "2", "MHz", "", "", ""),
    ("Recommended operating conditions", "Cooling plate temperature", "Tcoolingpad", "-40", "", "30", "°C", "", "", ""),
    ("Recommended operating conditions", "HalfSw threshold current", "Iout_halfsw", "600", "", "800", "mA", "", "", ""),
    ("Recommended operating conditions", "Inductor value", "L", "400", "", "500", "nH", "", "", ""),
    ("Recommended operating conditions", "Enable voltage", "Ven", "", "", "3.3", "V", "", "", ""),
    ("Recommended operating conditions", "Power Good voltage", "Vpgood", "", "", "3.3", "V", "", "", ""),
    # Electrical specifications (Table 2) - Power
    ("Power", "Input voltage supply range", "PVin, Vin", "5", "", "11", "V", "", "", ""),
    ("Power", "Input current (control, disabled)", "Iin", "", "2", "", "mA", "", "", ""),
    ("Power", "Output current (PCB in air)", "Iout", "", "", "1", "A", "", "", "f=1.8MHz, L=460nH, PCB in air"),
    ("Power", "Output current (active cooling)", "Iout", "", "", "4", "A", "", "", "f=1.8MHz, L=460nH, cooling plate 18°C"),
    ("Power", "Output power (PCB in air)", "Pout", "", "", "2", "W", "", "", "f=1.8MHz, L=460nH, PCB in air"),
    ("Power", "Output power (active cooling)", "Pout", "", "", "10", "W", "", "", "f=1.8MHz, L=460nH, cooling plate 18°C"),
    ("Power", "Power short input", "", "-10", "", "10", "mA", "M0001", "T0001", ""),
    ("Power", "Power short output", "", "-10", "", "10", "mA", "M0002", "T0002", ""),
    # PWM
    ("PWM", "Maximum Duty Cycle", "DMax", "", "100", "", "%", "", "", ""),
    ("PWM", "Minimum Duty Cycle", "DMin", "", "0", "", "%", "", "", ""),
    # Error Amplifier
    ("Error Amplifier", "DC Gain", "DCG", "", "90", "", "dB", "", "", "CL=1pF at VF Pin"),
    ("Error Amplifier", "Unity Gain-Bandwidth", "UGBW", "", "20", "", "MHz", "", "", "CL=1pF at VF Pin"),
    ("Error Amplifier", "Slew Rate", "SR", "", "10", "", "V/us", "", "", "CL=1pF at VF Pin"),
    # Under-Voltage Lockout
    ("Under-voltage lockout", "UVLO high threshold (Vin rising)", "VinStartTh", "", "4.79", "", "V", "M0013", "T0009", ""),
    ("Under-voltage lockout", "UVLO low threshold (Vin falling)", "VinStopTh", "", "4.52", "", "V", "M0014", "T0009", ""),
    ("Under-voltage lockout", "UVLO hysteresis", "UVLO_hyst", "0.10", "", "", "V", "M0037", "T0009", ""),
    # Enable
    ("Enable", "Enable start threshold", "EnStartTh", "", "815", "", "mV", "M0007", "T0006", ""),
    ("Enable", "Enable stop threshold", "EnStopTh", "", "730", "", "mV", "M0008", "T0006", ""),
    ("Enable", "Enable hysteresis", "Hyst", "0.05", "", "", "V", "M0035", "T0006", ""),
    ("Enable", "Enable pin series resistance", "EnSerRes", "", "10", "", "kΩ", "", "", "Limits current through ESD when not powered"),
    ("Enable", "Inv_Enable functionality", "", "Pass", "", "", "", "M0025", "T0013", "Functional pass/fail"),
    # Input leakage
    ("Input leakage", "Input leakage low (IIL)", "IIL", "-30", "", "", "uA", "M0005", "T0005", ""),
    ("Input leakage", "Input leakage high (IIH)", "IIH", "", "", "15", "uA", "M0006", "T0005", ""),
    # Protections
    ("Protections", "Over Current Protection peak level", "OCPpk", "", "6", "", "A", "", "", "Vin=10V, Vout=2.5V, f=1.8MHz, L=460nH"),
    ("Protections", "Over Current Protection average level", "OCPavg", "", "4.8", "", "A", "", "", "Vin=10V, Vout=2.5V, f=1.8MHz, L=460nH"),
    ("Protections", "Over Temperature start threshold", "OTPStartTh", "", "103", "", "°C", "", "", "Tj rising trip level"),
    ("Protections", "Over Temperature stop threshold", "OTPStopTh", "", "73", "", "°C", "", "", "Tj falling trip level"),
    # Soft Start
    ("Soft Start", "Soft Start duration", "SSt", "", "", "440", "us", "", "", "Vin=10V, Vout=2.5V, f=1.8MHz, L=410nH"),
    # Power Good
    ("Power Good", "Output Over Voltage PGood threshold", "OV", "", "+6.5", "", "%", "M0039", "T0010", ""),
    ("Power Good", "Output Under Voltage PGood threshold", "UV", "", "-6.5", "", "%", "M0038", "T0010", ""),
    ("Power Good", "PGood output Ron", "Ron", "100", "3500", "10000", "Ohm", "M0009", "T0007", ""),
    ("Power Good", "PGood output IOL", "IOL", "40", "50", "", "uA", "M0010", "T0007", ""),
    ("Power Good", "PGood output IOH", "IOH", "40", "50", "", "uA", "M0011", "T0007", ""),
    # PTAT
    ("PTAT", "PTAT temperature slope", "PTAT", "", "8.5", "", "mV/°C", "", "", "Converter disabled, environmental T sweep"),
    ("PTAT", "PTAT voltage (at 25°C)", "V_PTAT", "0.20", "", "0.80", "V", "M0022", "T0011", ""),
    # On-chip regulator
    ("On-chip regulator", "V33Dr load regulation", "K_50mA", "-10", "", "", "mV/mA", "M0012", "T0008", ""),
    ("On-chip regulator", "V33Dr voltage", "v33dr", "2.97", "", "3.66", "V", "M0036", "T0008", ""),
    ("On-chip regulator", "Bandgap reference voltage", "V_BGR", "0.57", "0.60", "0.63", "V", "M0015", "T0010", ""),
    # Output voltage monitor (PGood thresholds)
    ("Output voltage monitor (PGood thresholds)", "PGood rise threshold low", "VTH_PGOOD_RISE_L", "0.52", "0.561", "0.60", "V", "M0016", "T0010", ""),
    ("Output voltage monitor (PGood thresholds)", "PGood fall threshold low", "VTH_PGOOD_FALL_L", "0.52", "0.561", "0.60", "V", "M0017", "T0010", ""),
    ("Output voltage monitor (PGood thresholds)", "PGood rise threshold high", "VTH_PGOOD_RISE_H", "0.60", "0.639", "0.68", "V", "M0018", "T0010", ""),
    ("Output voltage monitor (PGood thresholds)", "PGood fall threshold high", "VTH_PGOOD_FALL_H", "0.60", "0.639", "0.68", "V", "M0019", "T0010", ""),
    # Startup behaviour
    ("Startup behaviour", "Startup output voltage", "V_start", "0.90", "", "", "V", "M0023", "T0012", ""),
    ("Startup behaviour", "Startup time", "T_start", "", "", "1.5", "ms", "M0024", "T0012", ""),
    # Line / Load regulation
    ("Line regulation", "Line regulation delta", "V_LINE_DELTA", "", "", "0.04", "V", "M0026", "T0014", ""),
    ("Load regulation", "Load regulation delta", "V_LOAD_DELTA", "0", "", "0.40", "V", "M0027", "T0015", ""),
    # Efficiency
    ("Efficiency", "Efficiency (at 0.5A load)", "EFFICIENCY", "35", "", "100", "%", "M0028", "T0016", ""),
    # Output ripple
    ("Output ripple", "Peak-to-peak output ripple", "Ripple", "", "", "0.1", "V", "", "T0017", ""),
    # Radiation hardness
    ("Radiation hardness", "Total Ionizing Dose (TID)", "", "> 1 MGy", "", "", "Si", "", "By design + lot acceptance", ""),
    ("Radiation hardness", "Single Event Effects (SEE)", "", "64", "", "", "MeV.cm²/mg", "", "Tested free of destructive SEEs", ""),
]

VIS_PARAMS = [
    ("Recommended operating conditions", "Power supply voltage", "V_SUPPLY_9V", "6", "9", "11", "V", "VER", ""),
    ("Recommended operating conditions", "Operation temperature (ambient)", "T_amb", "0", "", "50", "°C", "", ""),
    ("Recommended operating conditions", "Coaxial cable impedance", "Z_cable", "", "75", "", "Ohm", "", ""),
    ("Recommended operating conditions", "Max coaxial cable attenuation @ 625 MHz", "", "", "", "21.2", "dB", "", ""),
    ("Recommended operating conditions", "Max coaxial cable attenuation @ 30 MHz", "", "", "", "4.74", "dB", "", ""),
    ("Recommended operating conditions", "Recommended exposure time", "t_exp", "0", "", "5", "ms", "", ""),
    ("Electrical characteristics", "Module supply current", "I_SUPPLY_9V", "160", "", "257", "mA", "VER", ""),
    ("Electrical characteristics", "AC output impedance (HD-BNC)", "Z_out_AC", "", "75", "", "Ohm", "", ""),
    ("Integrated components", "Image sensor", "MAG-IMG002X1-NC", "", "", "", "", "", ""),
    ("Integrated components", "CoaXPress serializer", "MAG-CXP00002-NP", "", "", "", "", "", ""),
    ("Integrated components", "DC/DC converters", "2x MAG-PSU00001-NP", "", "", "", "", "", ""),
    ("Radiation hardness", "Total Ionizing Dose (TID)", "", "> 1 MGy", "", "", "Si", "By design + lot acceptance", ""),
    ("Radiation hardness", "Single Event Effects (SEE)", "", "By design", "", "", "", "By design", ""),
]


# ---------------------------------------------------------------------------
# GENERATE command
# ---------------------------------------------------------------------------
def generate_config():
    print("Reading check_limits files …")
    cxp_limits = read_check_limits(CXP_CHECK_LIMITS)
    psu_limits = read_check_limits(PSU_CHECK_LIMITS)
    print(f"  CXP: {len(cxp_limits)} Mxxxx entries")
    print(f"  PSU: {len(psu_limits)} Mxxxx entries")

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # --- IC product sheets ---
    img_limits = read_check_limits(IMG_CHECK_LIMITS)
    print(f"  IMG: {len(img_limits)} Mxxxx entries")

    for sheet_name, params, limits in [
        ("CXP00002", CXP_PARAMS, cxp_limits),
        ("IMG002X1", IMG_PARAMS, img_limits),
        ("PSU00001", PSU_PARAMS, psu_limits),
    ]:
        ws = wb.create_sheet(title=sheet_name)
        cols = IC_COLUMNS
        _write_ic_sheet(ws, cols, params, limits)

    # --- VIS module sheet ---
    ws = wb.create_sheet(title="VIS100xx")
    _write_vis_sheet(ws, VIS_COLUMNS, VIS_PARAMS)

    # --- Instructions sheet ---
    _write_instructions_sheet(wb)

    wb.save(CONFIG_XLSX)
    print(f"\nConfig file written to:\n  {CONFIG_XLSX}")
    print("\nOpen it in Excel, review/edit, then run:")
    print("  python chapter7_config_manager.py report")


def _write_ic_sheet(ws, cols, params, limits):
    # Header row
    for ci, (name, width, fill) in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=ci, value=name)
        cell.font = HEADER_FONT
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(ci)].width = width

    # TRUE/FALSE validation on Include column
    dv = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=False)
    dv.error = "Please enter TRUE or FALSE"
    dv.errorTitle = "Invalid value"
    ws.add_data_validation(dv)

    row_num = 2
    prev_section = None
    for entry in params:
        section, parameter, ds_label, ds_min, ds_typ, ds_max, unit, mxxxx, test_proc, comment = entry

        # Insert section header row if section changed
        if section != prev_section:
            ws.cell(row=row_num, column=1, value="TRUE").font = DATA_FONT
            ws.cell(row=row_num, column=2, value=section).font = SECTION_FONT
            for ci in range(1, len(cols) + 1):
                ws.cell(row=row_num, column=ci).fill = SECTION_FILL
                ws.cell(row=row_num, column=ci).border = THIN_BORDER
            dv.add(ws.cell(row=row_num, column=1))
            row_num += 1
            prev_section = section

        # Data row
        ws.cell(row=row_num, column=1, value="TRUE").font = DATA_FONT
        ws.cell(row=row_num, column=2, value=section).font = DATA_FONT
        ws.cell(row=row_num, column=3, value=parameter).font = DATA_FONT
        ws.cell(row=row_num, column=4, value=ds_label).font = DATA_FONT
        ws.cell(row=row_num, column=5, value=ds_min).font = DATA_FONT
        ws.cell(row=row_num, column=6, value=ds_typ).font = DATA_FONT
        ws.cell(row=row_num, column=7, value=ds_max).font = DATA_FONT
        ws.cell(row=row_num, column=8, value=unit).font = DATA_FONT
        ws.cell(row=row_num, column=9, value=mxxxx).font = DATA_FONT
        ws.cell(row=row_num, column=10, value=test_proc).font = DATA_FONT
        ws.cell(row=row_num, column=11, value="").font = DATA_FONT  # Test_Name (fill in Excel)

        # Auto-fill LSL/USL from check_limits if Mxxxx is known
        if mxxxx and mxxxx in limits:
            lsl = limits[mxxxx]["lsl"]
            usl = limits[mxxxx]["usl"]
            ws.cell(row=row_num, column=12, value=lsl if lsl is not None else "").font = DATA_FONT
            ws.cell(row=row_num, column=13, value=usl if usl is not None else "").font = DATA_FONT
        else:
            ws.cell(row=row_num, column=12, value="").font = DATA_FONT
            ws.cell(row=row_num, column=13, value="").font = DATA_FONT

        ws.cell(row=row_num, column=14, value=comment).font = DATA_FONT

        for ci in range(1, len(cols) + 1):
            ws.cell(row=row_num, column=ci).border = THIN_BORDER
        dv.add(ws.cell(row=row_num, column=1))
        row_num += 1

    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{row_num - 1}"
    ws.freeze_panes = "A2"


def _write_vis_sheet(ws, cols, params):
    for ci, (name, width, fill) in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=ci, value=name)
        cell.font = HEADER_FONT
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(ci)].width = width

    dv = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=False)
    ws.add_data_validation(dv)

    row_num = 2
    prev_section = None
    for entry in params:
        section, parameter, ds_label, ds_min, ds_typ, ds_max, unit, test_proc, comment = entry

        if section != prev_section:
            ws.cell(row=row_num, column=1, value="TRUE").font = DATA_FONT
            ws.cell(row=row_num, column=2, value=section).font = SECTION_FONT
            for ci in range(1, len(cols) + 1):
                ws.cell(row=row_num, column=ci).fill = SECTION_FILL
                ws.cell(row=row_num, column=ci).border = THIN_BORDER
            dv.add(ws.cell(row=row_num, column=1))
            row_num += 1
            prev_section = section

        ws.cell(row=row_num, column=1, value="TRUE").font = DATA_FONT
        ws.cell(row=row_num, column=2, value=section).font = DATA_FONT
        ws.cell(row=row_num, column=3, value=parameter).font = DATA_FONT
        ws.cell(row=row_num, column=4, value=ds_label).font = DATA_FONT
        ws.cell(row=row_num, column=5, value=ds_min).font = DATA_FONT
        ws.cell(row=row_num, column=6, value=ds_typ).font = DATA_FONT
        ws.cell(row=row_num, column=7, value=ds_max).font = DATA_FONT
        ws.cell(row=row_num, column=8, value=unit).font = DATA_FONT
        ws.cell(row=row_num, column=9, value=test_proc).font = DATA_FONT
        ws.cell(row=row_num, column=10, value="").font = DATA_FONT  # Test_Name (fill in Excel)
        ws.cell(row=row_num, column=11, value=comment).font = DATA_FONT

        for ci in range(1, len(cols) + 1):
            ws.cell(row=row_num, column=ci).border = THIN_BORDER
        dv.add(ws.cell(row=row_num, column=1))
        row_num += 1

    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{row_num - 1}"
    ws.freeze_panes = "A2"


def _write_instructions_sheet(wb):
    ws = wb.create_sheet(title="README", index=0)
    instructions = [
        "Chapter 7 Configuration File — README",
        "",
        "This workbook controls which parameters appear in Chapter 7 of the",
        "Vision Series Test Plan.  One sheet per product.",
        "",
        "COLUMN DESCRIPTIONS (IC sheets: CXP00002, IMG002X1, PSU00001):",
        "  Include        → TRUE to include this row in the report, FALSE to exclude",
        "  Section        → Category group (renders as bold header in the report table)",
        "  Parameter      → Human-readable parameter name shown in the report",
        "  Datasheet_Label→ Label as it appears in the product datasheet",
        "  DS_Min/Typ/Max → Values from the product datasheet (BLUE header)",
        "  Unit           → Measurement unit",
        "  Mxxxx          → Measurement ID in the check_limits database (editable!)",
        "  Test_Proc      → Test procedure ID (Txxxx or Functional/VER/etc.)",
        "  Test_Name      → Short human-readable test name (shown first in Chapter 7 report)",
        "  LSL            → Lower Specification Limit from check_limits (ORANGE header)",
        "  USL            → Upper Specification Limit from check_limits (ORANGE header)",
        "  Comment        → Free-text note shown in the report Comment column",
        "",
        "VIS100xx sheet has no Mxxxx/LSL/USL columns (module-level functional test).",
        "It includes Test_Proc and Test_Name like the IC sheets.",
        "",
        "HOW TO UPDATE:",
        "  1. Set Include=FALSE for any parameter you want to hide from the report.",
        "  2. Change the Mxxxx column to fix an incorrect measurement ID.",
        "  3. Edit LSL/USL if the auto-populated values need correction.",
        "  4. Add a Comment to explain why a parameter has no Mxxxx, is covered by",
        "     another measurement, or any other note you want in the report.",
        "  5. You can add new rows — just fill in all columns and set Include=TRUE.",
        "  6. You can reorder rows by cut-pasting — the section grouping is driven",
        "     by the Section column, not by row position.",
        "",
        "WHEN DONE EDITING:",
        "  Save this file, then run:",
        "    python chapter7_config_manager.py report",
        "  This will update Vision_Series_Test_Plan.md Chapter 7 and regenerate",
        "  the Word document.",
        "",
        "COLOR LEGEND:",
        "  Blue header columns  (DS_Min/Typ/Max) → values from the DATASHEET",
        "  Orange header columns (LSL/USL)        → values from CHECK_LIMITS tester DB",
        "  Grey rows                              → section headers (category grouping)",
    ]
    for ri, text in enumerate(instructions, start=1):
        cell = ws.cell(row=ri, column=1, value=text)
        cell.font = Font(name="Consolas", size=11)
        if ri == 1:
            cell.font = Font(name="Consolas", size=14, bold=True)
    ws.column_dimensions["A"].width = 90


# ---------------------------------------------------------------------------
# T0038 electro-optical test parameters (extra IMG table)
# ---------------------------------------------------------------------------
def _read_eo_descriptions(xlsx_path):
    """Read electro-optical parameter descriptions from the IMG xlsx.

    Returns {mid_str: {name, label, unit, definition}} where mid_str is e.g. 'M3800'.
    """
    descs = {}
    if not os.path.exists(xlsx_path):
        print(f"  [WARN] EO params file not found: {xlsx_path}")
        return descs
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["20251203 - list of parameters"]
    rows = list(ws.iter_rows(values_only=True))
    # Header at row index 3: MID, Name, Label, Unit, ..., Definition (col 10)
    for r in rows[4:]:
        mid_raw = str(r[0]).strip() if r[0] else ""
        if not mid_raw or not mid_raw.isdigit():
            continue
        mid = f"M{mid_raw}"
        name = str(r[1]).strip() if r[1] else ""
        label = str(r[2]).strip() if r[2] else ""
        unit = str(r[3]).strip() if r[3] else ""
        definition = str(r[10]).strip() if len(r) > 10 and r[10] else ""
        descs[mid] = {"name": name, "label": label, "unit": unit, "definition": definition}
    wb.close()
    return descs


def _read_t0038_params(check_limits_path):
    """Read all T0038 rows from check_limits.

    Returns list of dicts {mid, name_raw, lsl, usl} (keeps first occurrence per mid).
    """
    params = []
    seen = set()
    if not os.path.exists(check_limits_path):
        return params
    wb = load_workbook(check_limits_path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        wb.close()
        return params
    header = [str(c).strip() if c else "" for c in rows[0]]
    idx = {name: i for i, name in enumerate(header)}
    pid_col = idx.get("PARAMETER_ID", None)
    lsl_col = idx.get("LSL", None)
    usl_col = idx.get("USL", None)
    if pid_col is None:
        wb.close()
        return params
    mxxxx_re = re.compile(r'M\d{4}')
    for row in rows[1:]:
        pid = str(row[pid_col]) if row[pid_col] else ""
        if not pid.startswith("T0038"):
            continue
        match = mxxxx_re.search(pid)
        if not match:
            continue
        mid = match.group(0)
        if mid in seen:
            continue
        seen.add(mid)
        # Strip Txxxx_Mxxxx_ prefix to get the name
        name_raw = re.sub(r'^T\d{4}_M\d{4}_', '', pid)
        lsl_val = row[lsl_col] if lsl_col is not None else None
        usl_val = row[usl_col] if usl_col is not None else None
        params.append({"mid": mid, "name_raw": name_raw, "lsl": lsl_val, "usl": usl_val})
    wb.close()
    return params


def _build_t0038_markdown(t0038_params, eo_descs, existing_mids, unit_from_config):
    """Build markdown table for T0038 parameters not already in the config.

    existing_mids: set of Mxxxx strings already in the IMG config sheet.
    """
    filtered = [p for p in t0038_params if p["mid"] not in existing_mids]
    if not filtered:
        return ""

    lines = []
    lines.append("**Table 1b: MAG-IMG002X1-NC Electro-Optical Test Parameters (T0038)**")
    lines.append("")
    lines.append("| Name | Unit | LSL | USL | Description |")
    lines.append("| :--- | :--- | :--- | :--- | :--- |")

    for p in filtered:
        mid = p["mid"]
        desc_entry = eo_descs.get(mid, {})
        name = desc_entry.get("label", "") or p["name_raw"]
        unit = desc_entry.get("unit", "")
        definition = desc_entry.get("definition", "")
        # Clean up definition for markdown (no pipes)
        definition = definition.replace("|", "/").replace("\n", " ").strip()
        if definition in ("None", "nan"):
            definition = ""
        lsl = _fmt(p["lsl"])
        usl = _fmt(p["usl"])
        lines.append(f"| {name} | {unit} | {lsl} | {usl} | {definition} |")

    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Cross-reference tables (Chapter 6): test procedure → parameter names
# ---------------------------------------------------------------------------
def _build_xref_markdown(rows, product_name, product_key=""):
    """Build a cross-reference table mapping test procedures to measured parameters."""
    test_to_params = {}
    test_to_name = {}
    prod_labels = _PRODUCT_TEST_LABELS.get(product_key, {})
    for r in rows:
        tp = r.get("Test_Proc", "").strip()
        param = r.get("Parameter", "").strip()
        if not tp or tp in ("-", "None") or not param:
            continue
        explicit_name = (r.get("Test_Name", "") or "").strip()
        for t in re.split(r'[/,]', tp):
            t = t.strip()
            if not t or t in ("-", "None"):
                continue
            test_to_params.setdefault(t, []).append(param)
            if t not in test_to_name and explicit_name:
                test_to_name[t] = explicit_name

    if not test_to_params:
        return ""

    lines = []
    lines.append(f"**Parameter traceability: {product_name}**")
    lines.append("")
    lines.append("The table below maps each test procedure to the datasheet parameters it verifies. Parameters are listed by their Chapter 7 name.")
    lines.append("")
    lines.append("| Test procedure | Test ID | Parameters measured |")
    lines.append("| :--- | :--- | :--- |")

    for t in sorted(test_to_params.keys()):
        name = prod_labels.get(t) or test_to_name.get(t) or _fallback_test_name_from_proc(t)
        params_str = ", ".join(test_to_params[t])
        lines.append(f"| {name} | {t} | {params_str} |")

    return "\n".join(lines)


# Cross-reference section markers in Chapter 6
XREF_MARKERS = {
    "IMG002X1": ("#### 6.1.1 MAG-IMG002x1-NC", "#### 6.1.2"),
    "CXP00002": ("#### 6.1.2 MAG-CXP00002-NP", "#### 6.1.3"),
    "PSU00001": ("#### 6.1.3 MAG-PSU00001-NP", "### 6.2"),
}


# ---------------------------------------------------------------------------
# Appendix: Traceability matrix  (Option E — category × test ID)
# ---------------------------------------------------------------------------
APPENDIX_MARKER = "## Appendix A: Traceability Matrix"

PRODUCT_FULL_NAMES = {
    "IMG002X1": "MAG-IMG002X1-NC",
    "CXP00002": "MAG-CXP00002-NP",
    "PSU00001": "MAG-PSU00001-NP",
}


def _build_traceability_matrix(all_tables):
    """Build an appendix with one traceability matrix per IC product.

    Rows = parameter categories (Section), Columns = unique test IDs.
    Cells = X where at least one parameter in that category is tested by that ID.
    """
    lines = []
    lines.append(APPENDIX_MARKER)
    lines.append("")
    lines.append(
        "The matrices below provide a concise overview of which specification "
        "categories (rows) are verified by which test procedures (columns). "
        "An **X** indicates that at least one parameter in that category is "
        "measured by the test. Detailed parameter-level traceability is provided "
        "in the cross-reference tables of Chapter 6."
    )

    for product in ["IMG002X1", "CXP00002", "PSU00001"]:
        rows = all_tables.get(product)
        if not rows:
            continue

        # Collect unique test IDs (sorted) and section → set of test IDs
        section_tests = {}
        all_tests = set()
        for r in rows:
            section = r.get("Section", "").strip()
            tp = (r.get("Test_Proc", "") or "").strip()
            if not section or not tp or tp in ("-", "None"):
                continue
            for t in re.split(r'[/,]', tp):
                t = t.strip()
                if t and t not in ("-", "None"):
                    section_tests.setdefault(section, set()).add(t)
                    all_tests.add(t)

        if not all_tests:
            continue

        sorted_tests = sorted(all_tests)
        full_name = PRODUCT_FULL_NAMES.get(product, product)

        lines.append("")
        lines.append(f"### {full_name}")
        lines.append("")

        # Header
        header = "| Category |"
        sep = "| :--- |"
        for t in sorted_tests:
            header += f" {t} |"
            sep += " :---: |"
        lines.append(header)
        lines.append(sep)

        # Rows (one per section, ordered by first appearance)
        seen_sections = []
        for r in rows:
            s = r.get("Section", "").strip()
            if s and s not in seen_sections:
                seen_sections.append(s)

        for section in seen_sections:
            tests_in_section = section_tests.get(section, set())
            row_str = f"| {section} |"
            for t in sorted_tests:
                row_str += " X |" if t in tests_in_section else " |"
            lines.append(row_str)

    return "\n".join(lines)


def _update_appendix_in_markdown(content, appendix_md):
    """Insert or replace the traceability appendix at the end of the document."""
    pos = content.find(APPENDIX_MARKER)
    if pos != -1:
        content = content[:pos].rstrip() + "\n\n"
    else:
        content = content.rstrip() + "\n\n---\n\n"

    content += appendix_md + "\n"
    return content


def _update_xref_in_markdown(content, xref_tables):
    """Insert or update cross-reference tables at the end of each product's
    Chapter 6 section (before the next section marker)."""
    xref_marker = "**Parameter traceability:"

    for product, md_xref in xref_tables.items():
        if not md_xref or product not in XREF_MARKERS:
            continue

        start_marker, end_marker = XREF_MARKERS[product]
        start_idx = content.find(start_marker)
        if start_idx == -1:
            continue
        end_idx = content.find(end_marker, start_idx + len(start_marker))
        if end_idx == -1:
            end_idx = len(content)

        section = content[start_idx:end_idx]
        # Remove any existing xref table in this section
        xref_pos = section.find(xref_marker)
        if xref_pos != -1:
            section = section[:xref_pos].rstrip() + "\n\n"
        else:
            section = section.rstrip() + "\n\n"

        section += md_xref + "\n\n"
        content = content[:start_idx] + section + content[end_idx:]

    return content


# ---------------------------------------------------------------------------
# REPORT command  —  read config → markdown tables → update .md → build .docx
# ---------------------------------------------------------------------------
def generate_report():
    if not os.path.exists(CONFIG_XLSX):
        print(f"Config file not found: {CONFIG_XLSX}")
        print("Run 'python chapter7_config_manager.py generate' first.")
        sys.exit(1)

    # Re-read check_limits so LSL/USL are always fresh based on current Mxxxx
    print("Re-reading check_limits for fresh LSL/USL lookup …")
    fresh_limits = {
        "CXP00002": read_check_limits(CXP_CHECK_LIMITS),
        "PSU00001": read_check_limits(PSU_CHECK_LIMITS),
        "IMG002X1": read_check_limits(IMG_CHECK_LIMITS),
    }

    wb = load_workbook(CONFIG_XLSX, data_only=True)
    print("Reading config …")

    tables = {}
    for sheet_name in ["CXP00002", "IMG002X1", "PSU00001"]:
        if sheet_name not in wb.sheetnames:
            print(f"  [WARN] Sheet '{sheet_name}' not found — skipping")
            continue
        rows = _read_ic_sheet(wb[sheet_name])
        limits = fresh_limits.get(sheet_name, {})
        # Override LSL/USL from check_limits based on current Mxxxx in config
        for r in rows:
            mxxxx = r.get("Mxxxx", "").strip()
            if mxxxx and mxxxx in limits:
                lsl = limits[mxxxx].get("lsl")
                usl = limits[mxxxx].get("usl")
                r["LSL"] = str(lsl) if lsl is not None else ""
                r["USL"] = str(usl) if usl is not None else ""
            else:
                r["LSL"] = ""
                r["USL"] = ""
        tables[sheet_name] = rows

    if "VIS100xx" in wb.sheetnames:
        tables["VIS100xx"] = _read_vis_sheet(wb["VIS100xx"])

    wb.close()

    # Build markdown
    print("Generating markdown tables …")
    md_sections = {}

    if "IMG002X1" in tables:
        img_md = _build_ic_markdown(
            tables["IMG002X1"],
            "Table 1: MAG-IMG002X1-NC Key Specifications"
        )
        # Append T0038 electro-optical test parameters table
        existing_img_mids = {
            r.get("Mxxxx", "").strip()
            for r in tables["IMG002X1"]
            if r.get("Mxxxx", "").strip()
        }
        print("Building T0038 electro-optical test table …")
        t0038_params = _read_t0038_params(IMG_CHECK_LIMITS)
        eo_descs = _read_eo_descriptions(IMG_EO_PARAMS_XLSX)
        t0038_md = _build_t0038_markdown(t0038_params, eo_descs, existing_img_mids, {})
        if t0038_md:
            t0038_note = (
                "The table below lists the electro-optical and pixel-array defect test "
                "parameters measured during test T0038. The test limits shown are a first "
                "suggestion based on early characterisation data and are not yet final. "
                "Additional production data is needed to refine these limits so that they "
                "are both technically sound and economically viable."
            )
            img_md += "\n\n" + t0038_note + "\n\n" + t0038_md
            print(f"  {len([p for p in t0038_params if p['mid'] not in existing_img_mids])} T0038 parameters added")
        md_sections["IMG002X1"] = img_md

    if "CXP00002" in tables:
        md_sections["CXP00002"] = _build_ic_markdown(
            tables["CXP00002"],
            "Table 2: MAG-CXP00002-NP Key Specifications"
        )

    if "PSU00001" in tables:
        md_sections["PSU00001"] = _build_ic_markdown(
            tables["PSU00001"],
            "Table 3: MAG-PSU00001-NP Key Specifications"
        )

    if "VIS100xx" in tables:
        md_sections["VIS100xx"] = _build_vis_markdown(
            tables["VIS100xx"],
            "Table 4: MAG-VIS100xx-N Key Specifications"
        )

    _update_markdown(md_sections)

    # Build cross-reference tables for Chapter 6
    print("Generating cross-reference tables …")
    xref_tables = {}
    product_labels = {
        "IMG002X1": "MAG-IMG002X1-NC",
        "CXP00002": "MAG-CXP00002-NP",
        "PSU00001": "MAG-PSU00001-NP",
    }
    for product in ["IMG002X1", "CXP00002", "PSU00001"]:
        if product in tables:
            xref_tables[product] = _build_xref_markdown(
                tables[product], product_labels[product], product_key=product
            )

    with open(MD_FILE, "r", encoding="utf-8") as f:
        content = f.read()
    content = _update_xref_in_markdown(content, xref_tables)
    # Strip any leftover appendix from previous runs
    app_pos = content.find(APPENDIX_MARKER)
    if app_pos != -1:
        content = content[:app_pos].rstrip() + "\n"
    with open(MD_FILE, "w", encoding="utf-8") as f:
        f.write(content)

    print(f"Markdown updated: {MD_FILE}")

    # Build docx
    print("Generating Word document …")
    import subprocess
    convert_script = os.path.join(BASE_DIR, "convert_md_to_docx.py")
    result = subprocess.run(
        [sys.executable, convert_script],
        cwd=BASE_DIR, capture_output=True, text=True
    )
    if result.returncode == 0:
        print(result.stdout.strip())
    else:
        print(f"  [ERROR] docx conversion failed:\n{result.stderr}")


def _read_ic_sheet(ws):
    """Return list of row dicts from an IC product sheet."""
    rows_out = []
    all_rows = list(ws.iter_rows(values_only=True))
    if len(all_rows) < 2:
        return rows_out
    header = [str(c).strip() if c else "" for c in all_rows[0]]
    for row in all_rows[1:]:
        d = {}
        for hi, hname in enumerate(header):
            val = row[hi] if hi < len(row) else None
            d[hname] = str(val).strip() if val is not None else ""
        if d.get("Include", "").upper() != "TRUE":
            continue
        if d.get("Parameter", "") == "":
            continue
        rows_out.append(d)
    return rows_out


def _read_vis_sheet(ws):
    """Return list of row dicts from the VIS module sheet."""
    rows_out = []
    all_rows = list(ws.iter_rows(values_only=True))
    if len(all_rows) < 2:
        return rows_out
    header = [str(c).strip() if c else "" for c in all_rows[0]]
    for row in all_rows[1:]:
        d = {}
        for hi, hname in enumerate(header):
            val = row[hi] if hi < len(row) else None
            d[hname] = str(val).strip() if val is not None else ""
        if d.get("Include", "").upper() != "TRUE":
            continue
        if d.get("Parameter", "") == "":
            continue
        rows_out.append(d)
    return rows_out


def _fmt(val):
    """Format a cell value for markdown: strip 'None', trailing .0, etc."""
    if val in (None, "", "None", "nan", "NaN"):
        return "-"
    s = str(val).strip()
    if s == "":
        return "-"
    try:
        f = float(s)
        if f == int(f) and "." in s and not s.startswith(">"):
            s = str(int(f))
    except (ValueError, OverflowError):
        pass
    return s


# check_limits stores values in base SI (A, V, Hz, s, …).
# This map gives the multiplier to convert from base SI to the display unit.
_UNIT_SCALE = {
    "mA":    1e3,
    "uA":    1e6,
    "mV":    1e3,
    "V":     1,
    "MHz":   1e-6,
    "kHz":   1e-3,
    "Hz":    1,
    "Ohm":   1,
    "Ω":     1,
    "%":     1,
    "ms":    1e3,
    "us":    1e6,
    "s":     1,
    "mW":    1e3,
    "W":     1,
    "A":     1,
    "Gbps":  1,
    "Mbps":  1,
    "°C":    1,
    "mV/mA": 1,
}


# Per-product mapping from Txxxx to Chapter 6 test procedure name.
_PRODUCT_TEST_LABELS = {
    "IMG002X1": {
        "T0004": "VIL/VIH",
        "T0005": "VOL/VOH",
        "T0007": "IDDD",
        "T0010": "Static signal test",
        "T0024": "Frame modes",
        "T0038": "Optical test processing",
    },
    "CXP00002": {
        "T0004": "VIL/VIH",
        "T0005": "VOH/VOL",
        "T0006": "IOL/IOH",
        "T0007": "IDDD",
        "T0009": "Regulators",
        "T0011": "Regulators",
        "T0012": "Regulators",
        "T0013": "Oscillator",
        "T0020": "CXP Uplink Receiver Sensitivity",
        "T0021": "CXP Uplink Receiver Sensitivity",
        "T0022": "Internal Test Patterns",
        "T0023": "Camera Test",
    },
    "PSU00001": {
        "T0001": "Power Short",
        "T0002": "Load Short",
        "T0005": "Leakage",
        "T0006": "Input threshold",
        "T0007": "Output drive-strength",
        "T0008": "On-chip regulator",
        "T0009": "Under Voltage Lock Out",
        "T0010": "Output voltage monitor",
        "T0011": "Over Temperature Protection",
        "T0012": "Startup behaviour",
        "T0014": "Line regulation",
        "T0015": "Load regulation",
        "T0016": "Efficiency",
        "T0017": "Output voltage ripple",
    },
}

# Global fallback (used when product key is not available)
_TEST_ID_LABELS = {}
for _prod_labels in _PRODUCT_TEST_LABELS.values():
    for _k, _v in _prod_labels.items():
        if _k not in _TEST_ID_LABELS:
            _TEST_ID_LABELS[_k] = _v


def _fallback_test_name_from_proc(test_proc: str) -> str:
    """Derive a display name from Test_Proc when Test_Name is blank."""
    tp = (test_proc or "").strip()
    if not tp or tp in ("-", "None"):
        return ""
    if tp.lower() == "functional":
        return "Functional test"
    if tp == "VER":
        return "Module verification (VER)"
    if tp.startswith("By design"):
        return "Design assurance"
    parts_out = []
    for raw_tok in re.split(r"[/,]\s*", tp):
        tok = raw_tok.strip()
        if not tok:
            continue
        label = _TEST_ID_LABELS.get(tok)
        if label and label not in parts_out:
            parts_out.append(label)
        elif tok.lower() == "functional" and "Functional test" not in parts_out:
            parts_out.append("Functional test")
    if parts_out:
        return " + ".join(parts_out)
    return ""


def _chapter7_test_columns(row: dict) -> tuple:
    """Return (test_name, test_id) for Chapter 7 markdown tables."""
    raw_tp = (row.get("Test_Proc", "") or "").strip()
    explicit = (row.get("Test_Name", "") or "").strip()
    if explicit:
        return explicit, raw_tp
    if not raw_tp:
        return "", ""
    derived = _fallback_test_name_from_proc(raw_tp)
    return derived, raw_tp


def _convert_lsl_usl(raw_val, unit):
    """Convert a check_limits value (base SI) to the display unit."""
    if raw_val in (None, "", "None", "nan", "NaN"):
        return "-"
    s = str(raw_val).strip()
    if s == "" or s == "-":
        return "-"
    try:
        f = float(s)
    except (ValueError, OverflowError):
        return "-"
    scale = _UNIT_SCALE.get(unit, 1)
    converted = f * scale
    # Format nicely: drop unnecessary trailing zeros
    if converted == int(converted) and abs(converted) < 1e12:
        return str(int(converted))
    # Up to 6 significant figures, strip trailing zeros
    formatted = f"{converted:.6g}"
    return formatted


def _build_ic_markdown(rows, table_title):
    """Build a markdown table string for an IC product."""
    lines = []
    lines.append(f"**{table_title}**")
    lines.append("")
    lines.append("| Parameter | Datasheet Label | DS Min | DS Typ | DS Max | Unit | LSL | USL | Comment |")
    lines.append("| :--- | :--- | :--- | :--- | :--- | :--- | :--- | :--- | :--- |")

    prev_section = None
    for r in rows:
        section = r.get("Section", "")
        if section and section != prev_section:
            lines.append(f"| **{section}** | | | | | | | | |")
            prev_section = section

        param = r.get("Parameter", "-")
        ds_label = _fmt(r.get("Datasheet_Label", ""))
        ds_min = _fmt(r.get("DS_Min", ""))
        ds_typ = _fmt(r.get("DS_Typ", ""))
        ds_max = _fmt(r.get("DS_Max", ""))
        unit = _fmt(r.get("Unit", ""))
        lsl = _convert_lsl_usl(r.get("LSL", ""), unit)
        usl = _convert_lsl_usl(r.get("USL", ""), unit)
        comment = _fmt(r.get("Comment", ""))

        lines.append(
            f"| {param} | {ds_label} | {ds_min} | {ds_typ} | {ds_max} | {unit} | {lsl} | {usl} | {comment} |"
        )

    return "\n".join(lines)


def _build_vis_markdown(rows, table_title):
    """Build a markdown table string for the VIS module."""
    lines = []
    lines.append(f"**{table_title}**")
    lines.append("")
    lines.append("| Parameter | Datasheet Label | DS Min | DS Typ | DS Max | Unit | Comment |")
    lines.append("| :--- | :--- | :--- | :--- | :--- | :--- | :--- |")

    prev_section = None
    for r in rows:
        section = r.get("Section", "")
        if section and section != prev_section:
            lines.append(f"| **{section}** | | | | | | |")
            prev_section = section

        param = r.get("Parameter", "-")
        ds_label = _fmt(r.get("Datasheet_Label", ""))
        ds_min = _fmt(r.get("DS_Min", ""))
        ds_typ = _fmt(r.get("DS_Typ", ""))
        ds_max = _fmt(r.get("DS_Max", ""))
        unit = _fmt(r.get("Unit", ""))
        comment = _fmt(r.get("Comment", ""))

        lines.append(
            f"| {param} | {ds_label} | {ds_min} | {ds_typ} | {ds_max} | {unit} | {comment} |"
        )

    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Markdown file updater
# ---------------------------------------------------------------------------
SECTION_MARKERS = {
    "IMG002X1": ("### 7.1 MAG-IMG002X1-NC", "### 7.2"),
    "CXP00002": ("### 7.2 MAG-CXP00002-NP", "### 7.3"),
    "PSU00001": ("### 7.3 MAG-PSU00001-NP", "### 7.4"),
    "VIS100xx": ("### 7.4 MAG-VIS100xx-N", None),
}

TABLE_LIST_SECTION = "## 3. List of Tables"


def _update_markdown(md_sections):
    with open(MD_FILE, "r", encoding="utf-8") as f:
        content = f.read()

    for product, md_table in md_sections.items():
        start_marker, end_marker = SECTION_MARKERS[product]
        start_idx = content.find(start_marker)
        if start_idx == -1:
            print(f"  [WARN] Marker '{start_marker}' not found in markdown")
            continue

        if end_marker:
            end_idx = content.find(end_marker, start_idx + len(start_marker))
            if end_idx == -1:
                end_idx = len(content)
        else:
            end_idx = len(content)

        # Find the heading line (keep it) and replace everything until next section
        heading_end = content.index("\n", start_idx) + 1
        new_section = heading_end_content(start_marker, product, md_table)
        content = content[:heading_end] + new_section + content[end_idx:]

    # Update table of contents
    toc_start = content.find(TABLE_LIST_SECTION)
    if toc_start != -1:
        toc_line_end = content.index("\n", toc_start)
        next_section = content.find("\n## ", toc_line_end + 1)
        if next_section == -1:
            next_section = len(content)
        new_toc = _build_toc()
        content = content[:toc_line_end + 1] + new_toc + content[next_section:]

    with open(MD_FILE, "w", encoding="utf-8") as f:
        f.write(content)


def heading_end_content(heading, product, md_table):
    """Build the section body (after the ### heading line)."""
    product_titles = {
        "IMG002X1": "(Full HD CMOS Image Sensor)",
        "CXP00002": "(CoaXPress Serializer ASIC)",
        "PSU00001": "(10W Synchronous Step-Down DC/DC Converter)",
        "VIS100xx": "(Electronic Vision Module)",
    }
    lines = [
        "",
        md_table,
        "",
    ]
    return "\n".join(lines) + "\n"


def _build_toc():
    return """
- Table 1: MAG-IMG002X1-NC Key Specifications
- Table 1b: MAG-IMG002X1-NC Electro-Optical Test Parameters (T0038)
- Table 2: MAG-CXP00002-NP Key Specifications
- Table 3: MAG-PSU00001-NP Key Specifications
- Table 4: MAG-VIS100xx-N Key Specifications
"""


# ---------------------------------------------------------------------------
# UPDATE command — add missing seed parameters to existing config
# ---------------------------------------------------------------------------
def update_config():
    """Add parameters from seed data that are not yet in the config xlsx."""
    if not os.path.exists(CONFIG_XLSX):
        print("No existing config found — running full generate instead.")
        generate_config()
        return

    print("Reading check_limits files …")
    cxp_limits = read_check_limits(CXP_CHECK_LIMITS)
    psu_limits = read_check_limits(PSU_CHECK_LIMITS)

    wb = load_workbook(CONFIG_XLSX)

    img_limits = read_check_limits(IMG_CHECK_LIMITS)

    all_seeds = {
        "CXP00002": (CXP_PARAMS, cxp_limits),
        "IMG002X1": (IMG_PARAMS, img_limits),
        "PSU00001": (PSU_PARAMS, psu_limits),
    }

    for sheet_name, (seed_params, limits) in all_seeds.items():
        if sheet_name not in wb.sheetnames:
            print(f"  [WARN] Sheet '{sheet_name}' not found — skipping")
            continue
        ws = wb[sheet_name]
        _update_ic_sheet(ws, seed_params, limits)

    wb.save(CONFIG_XLSX)
    print(f"\nConfig updated: {CONFIG_XLSX}")
    print("Review the new rows (marked Include=TRUE), then run:")
    print("  python chapter7_config_manager.py report")


def _update_ic_sheet(ws, seed_params, limits):
    """Append seed rows whose Parameter name is not yet in the sheet."""
    all_rows = list(ws.iter_rows(min_row=2, values_only=True))
    existing_params = set()
    for row in all_rows:
        param_val = str(row[2]).strip() if row[2] else ""
        if param_val:
            existing_params.add(param_val)

    last_row = ws.max_row
    added = 0
    prev_section = None
    for entry in seed_params:
        section, parameter, ds_label, ds_min, ds_typ, ds_max, unit, mxxxx, test_proc, comment = entry
        if parameter in existing_params:
            continue

        # Insert section header if this is a new section not yet in the sheet
        if section != prev_section:
            section_exists = any(
                str(r[1]).strip() == section for r in all_rows if r[1]
            )
            if not section_exists:
                last_row += 1
                ws.cell(row=last_row, column=1, value="TRUE").font = DATA_FONT
                ws.cell(row=last_row, column=2, value=section).font = SECTION_FONT
                for ci in range(1, len(IC_COLUMNS) + 1):
                    ws.cell(row=last_row, column=ci).fill = SECTION_FILL
                    ws.cell(row=last_row, column=ci).border = THIN_BORDER
            prev_section = section

        last_row += 1
        ws.cell(row=last_row, column=1, value="TRUE").font = DATA_FONT
        ws.cell(row=last_row, column=2, value=section).font = DATA_FONT
        ws.cell(row=last_row, column=3, value=parameter).font = DATA_FONT
        ws.cell(row=last_row, column=4, value=ds_label).font = DATA_FONT
        ws.cell(row=last_row, column=5, value=ds_min).font = DATA_FONT
        ws.cell(row=last_row, column=6, value=ds_typ).font = DATA_FONT
        ws.cell(row=last_row, column=7, value=ds_max).font = DATA_FONT
        ws.cell(row=last_row, column=8, value=unit).font = DATA_FONT
        ws.cell(row=last_row, column=9, value=mxxxx).font = DATA_FONT
        ws.cell(row=last_row, column=10, value=test_proc).font = DATA_FONT
        ws.cell(row=last_row, column=11, value="").font = DATA_FONT  # Test_Name
        if mxxxx and mxxxx in limits:
            ws.cell(row=last_row, column=12, value=limits[mxxxx]["lsl"]).font = DATA_FONT
            ws.cell(row=last_row, column=13, value=limits[mxxxx]["usl"]).font = DATA_FONT
        else:
            ws.cell(row=last_row, column=12, value="").font = DATA_FONT
            ws.cell(row=last_row, column=13, value="").font = DATA_FONT
        ws.cell(row=last_row, column=14, value=comment).font = DATA_FONT
        for ci in range(1, len(IC_COLUMNS) + 1):
            ws.cell(row=last_row, column=ci).border = THIN_BORDER
        added += 1
        existing_params.add(parameter)

    print(f"  {ws.title}: added {added} new rows")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    if len(sys.argv) < 2 or sys.argv[1] not in ("generate", "report", "update"):
        print("Usage:")
        print("  python chapter7_config_manager.py generate   # create config xlsx from scratch")
        print("  python chapter7_config_manager.py update     # add missing params to existing config")
        print("  python chapter7_config_manager.py report      # read config → update md + docx")
        sys.exit(1)

    if sys.argv[1] == "generate":
        generate_config()
    elif sys.argv[1] == "update":
        update_config()
    elif sys.argv[1] == "report":
        generate_report()
