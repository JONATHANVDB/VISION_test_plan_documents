"""Generate Vision Series Compliance Matrix Excel file.

Source: Projects/VISION_test_plan_documents/Vision_Series_Production_Test_Plan.md
Output: Projects/VISION_test_plan_documents/Vision_Series_Compliance_Matrix.xlsx
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# =========================
# Settings (edit defaults)
# =========================
OUTPUT_DIR = Path(__file__).resolve().parent
OUTPUT_FILE = OUTPUT_DIR / "Vision_Series_Compliance_Matrix.xlsx"

# Column headers for parameter sheets
HEADERS = [
    "Parameter",
    "Datasheet Label",
    "Description",
    "Class",
    "Test Level",
    "Test ID",
    "Test Procedure",
    "LSL",
    "USL",
    "LTL",
    "UTL",
    "Unit",
    "Bin (fail)",
    "Comment",
]

# Class: External = in datasheet Chapter 5, Internal = monitored but not in datasheet
# Test level codes: D = By design, S = Simulation, C = Characterization, P = Production test

# Bin definitions applied across products
BIN_LIST = [
    (1, "PROD",   "Pass",   "Part passes all production test limits (LTL/UTL) and is sellable as production grade."),
    (2, "EM",     "Accept", "Part is outside PROD limits but within Engineering Model tolerance; re-binned as EM grade (applies to selected pixel-array / DSNU / PRNU / hot-pixel parameters on MAG-IMG002x1-NC)."),
    (3, "VI",     "Fail",   "Part fails Visual Inspection stage (cracks, delamination, contamination, bond-wire or solder anomalies)."),
    (4, "CONT",   "Fail",   "Part fails Continuity / Power-Short / Load-Short gate (opens, shorts, ESD-diode anomalies)."),
    (5, "PARAM",  "Fail",   "Part fails a standard parametric production test (LTL/UTL exceeded) on an electrical / analog parameter."),
    (6, "FUNC",   "Fail",   "Part fails a functional test (SPI, CXP, PLL lock, testpattern, trigger, OTP retention, image capture)."),
    (7, "OPT",    "Fail",   "Part fails an electro-optical / pixel-array test (T0035 defect map or T0038 PTC/DC/SN)."),
    (9, "REJECT", "Fail",   "Part is outside even the EM acceptance range; part is rejected / scrapped."),
]

# =========================
# Helper constants
# =========================
EXT = "External"   # parameter is published in datasheet Chapter 5
INT = "Internal"   # parameter is monitored internally only

# Short test-level groups
D = "D"
S = "S"
C = "C"
P = "P"


def _row(parameter, ds_label, desc, klass, level, tid, proc, lsl, usl, ltl, utl,
         unit, bin_fail, comment=""):
    """Helper to build a row as a dict matching HEADERS."""
    return [parameter, ds_label, desc, klass, level, tid, proc, lsl, usl, ltl,
            utl, unit, bin_fail, comment]


# =========================
# MAG-IMG002x1-NC (Image Sensor IC)
# =========================
IMG_ROWS = [
    # ---- Current consumption (External) ----
    _row("VDDD current (active, FHD10 10-bit)", "I_VDDD",
         "Digital supply current in active mode, FHD10 10-bit video mode. "
         "Measured by IDDD test procedure in all FT stages.",
         EXT, P, "T0007", "IDDD",
         "-", 1000, 20, 300, "mA", 5,
         "Max 1000 mA only at startup/reset."),
    _row("VDDA+VDDIO+VD_RST current (active, FHD10 10-bit)", "I_VDDA (combined)",
         "Combined analog + IO + pixel reset supply current in active mode, FHD10 10-bit. "
         "Split over 3 PSU channels; the stated LTL/UTL apply to the VDDA channel.",
         EXT, P, "T0007", "IDDD",
         "-", "-", 20, 150, "mA", 5,
         "Split over 3 PSU channels; LSL/USL of VDDA is given here."),

    # ---- Clocking and control signal thresholds ----
    _row("External clock reference", "REFCLK",
         "External reference clock frequency range accepted by the device.",
         EXT, D, "-", "-",
         10, 40, "-", "-", "MHz", "-",
         "Fixed to 20MHz in test setup hardware."),
    _row("TRIG_IN input threshold low", "TRIG_in_l",
         "Falling-edge input low threshold of the TRIG_IN pin. Pin is rising-edge triggered; "
         "low threshold characterised only.",
         EXT, C, "T0004", "VIL/VIH",
         0.74, 1.4, "-", "-", "V", "-",
         "Pin is of type: rising edge triggered."),
    _row("TRIG_IN input threshold high", "TRIG_in_h",
         "Rising-edge input high threshold of the TRIG_IN pin.",
         EXT, P, "T0004", "VIL/VIH",
         1.56, 2.39, 1.6, 2.6, "V", 5,
         ""),
    _row("TRIG_IN input hysteresis", "TRIG_in_hyst",
         "Hysteresis between TRIG_IN high and low thresholds.",
         EXT, C, "T0004", "VIL/VIH",
         0.66, 1.15, "-", "-", "V", "-",
         "Pin is of type: rising edge triggered."),
    _row("RSTB input threshold (no hysteresis)", "RSTB_th_lvl",
         "Reset-bar input threshold level. No hysteresis is specified on this pin.",
         EXT, P, "T0004", "VIL/VIH",
         "-", "-", 1.4, 2.3, "V", 5,
         ""),

    # ---- SPI interface ----
    _row("SPI master clock speed", "SPI_CLK",
         "Maximum SPI master clock frequency supported by the device.",
         EXT, D, "-", "-",
         "-", 20, "-", "-", "MHz", "-",
         "Fixed to 20MHz in test setup hardware."),
    _row("SPI MISO output voltage (no load)", "SPI_out",
         "SPI MISO output high voltage. Measured at 1 mA load.",
         EXT, P, "T0005", "VOL/VOH",
         2.97, 3.63, 2.7, "-", "V", 5,
         "Tested with 1 mA load."),
    _row("SPI input threshold low", "SPI_in_l",
         "SPI MOSI input low threshold.",
         EXT, P, "T0004", "VIL/VIH",
         0.74, 1.4, 0.6, 1.7, "V", 5,
         "s_mosi_vil (representative)."),
    _row("SPI input threshold high", "SPI_in_h",
         "SPI MOSI input high threshold.",
         EXT, P, "T0004", "VIL/VIH",
         1.56, 2.39, 1.3, 2.3, "V", 5,
         "s_mosi_vih (representative)."),
    _row("SPI input hysteresis", "SPI_in_hyst",
         "Hysteresis between SPI MOSI high and low thresholds.",
         EXT, P, "T0004", "VIL/VIH",
         0.66, 1.15, 0.3, 0.9, "V", 5,
         "s_mosi_hysteresis (representative)."),

    # ---- Digital video output interface ----
    _row("VOH (no resistive load)", "VOH",
         "Digital video output high voltage. Measured at 1 mA load.",
         EXT, P, "T0005", "VOL/VOH",
         2.97, 3.63, 2.7, "-", "V", 5,
         "Tested with 1 mA load."),
    _row("VOL (no resistive load)", "VOL",
         "Digital video output low voltage. Measured at 1 mA load.",
         EXT, P, "T0005", "VOL/VOH",
         "-", "-", "-", 0.3, "V", 5,
         "Tested with 1 mA load."),
    _row("Pixel output clock speed", "PIXCLK_speed",
         "Pixel output clock speed. Only characterised at full speed (FHD30); production "
         "test runs FHD10/HD40/VGA50 which exercises half the maximum speed.",
         EXT, C, "T0024", "Frame modes",
         50, 100, "pass/fail", "-", "MHz", 6,
         "Only tested at speed during characterization (FHD30 mode)."),
    _row("Data output line speed", "DV_X_speed",
         "Parallel data output line speed. Same note as PIXCLK_speed.",
         EXT, C, "T0024", "Frame modes",
         "-", 50, "pass/fail", "-", "MHz", 6,
         "Only tested at speed during characterization (FHD30 mode)."),

    # ---- Analog signals ----
    _row("Bandgap reference voltage (VDD12BG)", "Vvdd12bg",
         "Internal bandgap reference voltage exposed through the test mux.",
         EXT, P, "T0010", "Static signal test",
         1.14, 1.26, 1.05, 1.35, "V", 5, ""),

    # ---- Electro-optical characteristics (External / datasheet) ----
    _row("Effective number of pixels", "-",
         "Effective pixel array resolution (H x V).",
         EXT, D, "-", "-", "-", "-", "-", "-", "1920 (H) x 1080 (V)", "-",
         "Guaranteed by design."),
    _row("Pixel size", "-",
         "Pixel pitch in micrometres.",
         EXT, D, "-", "-", "-", "-", "-", "-", "5 um x 5 um", "-",
         "Guaranteed by design."),
    _row("Number of black rows", "-",
         "Number of optical black reference rows in the array.",
         EXT, D, "-", "-", "-", "-", "-", "-", "64", "-",
         "Guaranteed by design."),
    _row("Shutter type", "-",
         "Sensor shutter mechanism.",
         EXT, D, "-", "-", "-", "-", "-", "-", "Rolling shutter", "-",
         "Guaranteed by design."),
    _row("Full well charge", "-",
         "Maximum charge that can be accumulated per pixel before saturation.",
         EXT, C, "-", "-", "-", "-", "-", "-", "e-", "-",
         "Design target 94000 e-, related to saturation limit; only tested during characterization."),
    _row("Conversion gain", "-",
         "Charge-to-voltage conversion gain of the pixel + readout chain.",
         EXT, P, "T0038", "Optical test processing",
         "-", "-", 0.006, 0.018, "mV/e-", 7, ""),
    _row("Responsivity", "-",
         "Ratio between pixel output code and incident photons.",
         EXT, P, "T0038", "Optical test processing",
         "-", "-", 0.02895, 0.04825, "DN/p", 7, ""),
    _row("Temporal noise", "-",
         "RMS temporal (read) noise at the pixel output referred to electrons.",
         EXT, P, "T0038", "Optical test processing",
         "-", "-", "-", 120, "e-", 7,
         "Datasheet typical: 66 e-."),
    _row("Dynamic range", "-",
         "Ratio between saturation and noise floor, expressed in dB.",
         EXT, P, "T0038", "Optical test processing",
         "-", "-", 57, 65, "dB", 7,
         "Datasheet typical: 62 dB."),
    _row("SNR_MAX", "-",
         "Maximum signal-to-noise ratio at the saturation level.",
         EXT, P, "T0038", "Optical test processing",
         "-", "-", 46, 54, "dB", 7,
         "Datasheet typical: 49.5 dB."),
    _row("Dark current (DC)", "-",
         "Mean dark current generated per pixel per second.",
         EXT, P, "T0038", "Optical test processing",
         "-", "-", "-", 2500, "e-/s/px", 7, ""),
    _row("Dark current non uniformity (DCNU)", "-",
         "Spread of dark current values across the pixel array.",
         EXT, P, "T0038", "Optical test processing",
         "-", "-", "-", 4000, "e-/s/px", 7, ""),
    _row("Dark signal non uniformity (DSNU)", "-",
         "Pixel-to-pixel offset variation under dark conditions.",
         EXT, P, "T0038", "Optical test processing",
         "-", "-", "-", 600, "e-", 7,
         "Datasheet typical: 587.2 e-."),
    _row("Photo response non uniformity (PRNU)", "-",
         "Pixel-to-pixel gain variation under uniform illumination.",
         EXT, P, "T0038", "Optical test processing",
         "-", 2.5, -2.5, 2.5, "% of mean", 7, ""),
    _row("Color filters", "-",
         "Colour filter array type.",
         EXT, D, "-", "-", "-", "-", "-", "-", "BayerRG (RGGB)", "-",
         "Guaranteed by design."),
    _row("Programmable features", "-",
         "Programmable sensor parameters: exposure time, sub-sampling, X-Y mirroring, ADC resolution.",
         EXT, C, "-", "-", "-", "-", "-", "-", "-", "-",
         "Features are tested during characterization."),
    _row("ADC resolution", "-",
         "On-chip ADC bit-depth options.",
         EXT, D, "-", "-", "-", "-", "-", "-", "10 and 12 bit", "-",
         "Guaranteed by design."),
    _row("Interface", "-",
         "Image data output interface type.",
         EXT, D, "-", "-", "-", "-", "-", "-", "10/12b parallel", "-",
         "Fixed to 10b for all production tests."),
    _row("Cover glass lid", "-",
         "Sensor package cover glass material.",
         EXT, D, "-", "-", "-", "-", "-", "-", "Corning 7980 0F", "-",
         "Guaranteed by design."),

    # ---- T0035 pixel-array defect parameters (Internal) ----
    _row("Number of dead pixels", "-",
         "Dead pixel = pixel with unusually low (positive or negative) response to light "
         "(stuck-at / negative-gain). Dead pixels in dead rows/columns are excluded.",
         INT, P, "T0035", "Optical defect map",
         "-", "-", "-", 50, "px", 7, ""),
    _row("Number of clusters with 2 dead pixels", "-",
         "Count of 2-neighbour dead pixel clusters on the same channel.",
         INT, P, "T0035", "Optical defect map",
         "-", "-", "-", 50, "cl", 7, ""),
    _row("Number of clusters with >3 dead pixels", "-",
         "Count of clusters with 3 or more neighbour dead pixels on the same channel.",
         INT, P, "T0035", "Optical defect map",
         "-", "-", "-", 50, "cl", 7, ""),
    _row("Number of dead rows", "-",
         "Row with 100 or more dead pixels.",
         INT, P, "T0035", "Optical defect map",
         "-", "-", "-", 5, "rows", 7, ""),
    _row("Number of dead columns", "-",
         "Column with 100 or more dead pixels.",
         INT, P, "T0035", "Optical defect map",
         "-", "-", "-", 5, "cols", 7, ""),
    _row("Number of bright pixels", "-",
         "Pixel with unusually high value in dark (outlier in DC histogram, broken pixel "
         "or high-offset pixel). Bright pixels in bright rows/columns are excluded.",
         INT, P, "T0035", "Optical defect map",
         "-", "-", "-", 50, "px", 7, ""),
    _row("Number of clusters with 2 bright pixels", "-",
         "Count of 2-neighbour bright pixel clusters on the same channel.",
         INT, P, "T0035", "Optical defect map",
         "-", "-", "-", 50, "cl", 7, ""),
    _row("Number of clusters with >3 bright pixels", "-",
         "Count of clusters with 3 or more neighbour bright pixels on the same channel.",
         INT, P, "T0035", "Optical defect map",
         "-", "-", "-", 50, "cl", 7, ""),
    _row("Number of bright rows", "-",
         "Row with 100 or more bright pixels.",
         INT, P, "T0035", "Optical defect map",
         "-", "-", "-", 5, "rows", 7, ""),
    _row("Number of bright cols", "-",
         "Column with 100 or more bright pixels.",
         INT, P, "T0035", "Optical defect map",
         "-", "-", "-", 5, "cols", 7, ""),
    _row("Number of bad columns", "-",
         "bad_col = dead_col + bright_col. PROD: USL=0; EM: 1-5 bad columns; Reject: >5.",
         INT, P, "T0035", "Optical defect map",
         "-", "-", "-", 0, "cols", 2,
         "Bin 2 (EM) if 1-5, Bin 9 (Reject) if >5."),
    _row("Number of bad rows", "-",
         "bad_row = dead_row + bright_row. PROD: USL=0; EM: 1-5 bad rows; Reject: >5.",
         INT, P, "T0035", "Optical defect map",
         "-", "-", "-", 0, "rows", 2,
         "Bin 2 (EM) if 1-5, Bin 9 (Reject) if >5."),
    _row("Distance between bad columns", "-",
         "Minimum column separation between bad columns.",
         INT, P, "T0035", "Optical defect map",
         "-", "-", 5, "-", "cols", 7, ""),
    _row("Distance between bad rows", "-",
         "Minimum row separation between bad rows.",
         INT, P, "T0035", "Optical defect map",
         "-", "-", 5, "-", "rows", 7, ""),

    # ---- T0038 electro-optical internal parameters ----
    _row("Typical system gain", "-",
         "System gain = conversion gain / ADC gain (typical).",
         INT, P, "T0038", "Optical test processing",
         "-", "-", 0.022, 0.066, "DN/e-", 7, ""),
    _row("DSNU local stdev", "-",
         "Offset variation between pixels in close proximity.",
         INT, P, "T0038", "Optical test processing",
         "-", "-", "-", 1500, "e-", 7, ""),
    _row("Column FPN (no black-row correction) global stdev", "-",
         "DSNU-like metric without column FPN reduction; dominated by column contribution.",
         INT, P, "T0038", "Optical test processing",
         "-", "-", "-", "-", "DN", 7, ""),
    _row("Column FPN (no black-row correction) local stdev", "-",
         "Local column FPN spread without black-row correction.",
         INT, P, "T0038", "Optical test processing",
         "-", "-", "-", "-", "DN", 7, ""),
    _row("Number of DSNU global outliers", "-",
         "Outliers in the offset histogram. PROD: USL=50; EM: up to 150; Reject: >150.",
         INT, P, "T0038", "Optical test processing",
         "-", "-", "-", 50, "px", 2,
         "Bin 2 (EM) up to 150, Bin 9 (Reject) >150."),
    _row("DSNU global stdev, row contribution", "-",
         "Row-wise component of global DSNU variation.",
         INT, P, "T0038", "Optical test processing",
         "-", "-", "-", 600, "e-", 7, ""),
    _row("DSNU global stdev, col contribution", "-",
         "Column-wise component of global DSNU variation.",
         INT, P, "T0038", "Optical test processing",
         "-", "-", "-", 600, "e-", 7, ""),
    _row("Number of PRNU global outliers", "-",
         "Outliers in the gain histogram. PROD: USL=150; EM: up to 1000; Reject: >1000.",
         INT, P, "T0038", "Optical test processing",
         "-", "-", "-", 150, "px", 2,
         "Bin 2 (EM) up to 1000, Bin 9 (Reject) >1000."),
    _row("PRNU global stdev, row contribution", "-",
         "Row-wise component of global PRNU variation.",
         INT, P, "T0038", "Optical test processing",
         "-", "-", -2.5, 2.5, "% of mean", 7, ""),
    _row("PRNU global stdev, col contribution", "-",
         "Column-wise component of global PRNU variation.",
         INT, P, "T0038", "Optical test processing",
         "-", "-", -2.5, 2.5, "% of mean", 7, ""),
    _row("Hot pixels > criteria 1 (>5x avg)", "-",
         "Tail of DC histogram, criterion 1. PROD: USL=1200; EM: up to 5000; Reject: >5000.",
         INT, P, "T0038", "Optical test processing",
         "-", "-", "-", 1200, "px", 2,
         "Bin 2 (EM) up to 5000, Bin 9 (Reject) >5000."),
    _row("Hot pixels > criteria 2 (>10x avg)", "-",
         "Tail of DC histogram, criterion 2. PROD: USL=600; EM: up to 2500; Reject: >2500.",
         INT, P, "T0038", "Optical test processing",
         "-", "-", "-", 600, "px", 2,
         "Bin 2 (EM) up to 2500, Bin 9 (Reject) >2500."),
    _row("Hot pixels > criteria 3 (>100x avg)", "-",
         "Tail of DC histogram, criterion 3. PROD: USL=5; EM: up to 20; Reject: >20.",
         INT, P, "T0038", "Optical test processing",
         "-", "-", "-", 5, "px", 2,
         "Bin 2 (EM) up to 20, Bin 9 (Reject) >20."),
    _row("Percentage of clusters with 2 hot pixels", "-",
         "Cluster size of 2 px; crosstalk causes hot pixels to appear in groups of 2.",
         INT, P, "T0038", "Optical test processing",
         "-", "-", 0, 100, "%", 7, ""),
    _row("Percentage of clusters with >3 hot pixels", "-",
         "Larger clusters help distinguish crosstalk from other defect mechanisms.",
         INT, P, "T0038", "Optical test processing",
         "-", "-", 0, 100, "%", 7, ""),
    _row("PRNU local stdev", "-",
         "Gain (light response) variation between pixels in close proximity.",
         INT, P, "T0038", "Optical test processing",
         "-", "-", -1, 1.5, "% of mean", 7, ""),
    _row("Number of DSNU local outliers", "-",
         "Local DSNU histogram outliers.",
         INT, P, "T0038", "Optical test processing",
         "-", "-", "-", 50, "px", 7, ""),
    _row("Number of column offset outliers (no black correction)", "-",
         "Columns with outlier offset values without black-row correction.",
         INT, P, "T0038", "Optical test processing",
         "-", "-", "-", 50, "col", 7, ""),
    _row("Number of PRNU local outliers", "-",
         "Local PRNU histogram outliers.",
         INT, P, "T0038", "Optical test processing",
         "-", "-", "-", 50, "px", 7, ""),

    # ---- Other internal functional / parametric checks ----
    _row("Pin continuity (all pins)", "-",
         "ESD-diode voltage drop check on every pin with supplies off; guards against opens.",
         INT, P, "T0001", "Continuity",
         "-", "-", "-", "-", "V", 4, ""),
    _row("Power short (per power rail)", "-",
         "Low-voltage short-circuit check on each supply domain.",
         INT, P, "T0002", "Power Short",
         "-", "-", "-", "-", "mA", 4, ""),
    _row("Pin leakage", "-",
         "Leakage current on input / tristate output pins.",
         INT, P, "T0003", "Leakage",
         "-", "-", "-", "-", "nA", 5, ""),
    _row("IOL/IOH", "-",
         "Digital output sink / source current capability.",
         INT, P, "T0006", "IOL/IOH",
         "-", "-", "-", "-", "mA", 5, ""),
    _row("Test buffer offset", "-",
         "Offset of both on-die test buffers on the test mux.",
         INT, P, "T0008", "Test buffer offset",
         "-", "-", "-", "-", "mV", 5, ""),
    _row("PLL open-loop VCO frequency", "-",
         "VCO open-loop frequency vs forced v_tune voltage.",
         INT, P, "T0014", "PLL open-loop",
         "-", "-", "-", "-", "MHz", 5, ""),
    _row("PLL lock", "-",
         "PLL lock status and locked-frequency check.",
         INT, P, "T0015", "PLL lock",
         "-", "-", "-", "-", "-", 6, ""),
    _row("Ramp generator delay", "-",
         "Readout ramp-generator propagation delay.",
         INT, P, "T0016", "Ramp gen test",
         "-", "-", "-", "-", "ns", 5, ""),
    _row("Power-on reset (POR)", "-",
         "Validates power-on reset behaviour of the device.",
         INT, P, "T0017", "POR",
         "-", "-", "-", "-", "-", 6, ""),
    _row("Trigger mode", "-",
         "Response to external trigger pulse.",
         INT, P, "T0018", "Trigger mode",
         "-", "-", "-", "-", "-", 6, ""),
    _row("Testpattern generator", "-",
         "Internal test-pattern capture compared against reference.",
         INT, P, "T0019", "Testpattern generator",
         "-", "-", "-", "-", "-", 6, ""),
    _row("Serial interface (SPI)", "-",
         "SPI functional read/write verification.",
         INT, P, "T0020", "Serial interface",
         "-", "-", "-", "-", "-", 6, ""),
    _row("Block trim values", "-",
         "Best-fit trim code per internal block (voltage ref, current ref, ramp gen).",
         INT, P, "T0021", "Trim test / Ramp gen trim",
         "-", "-", "-", "-", "code", 5,
         "Performed in TRIM stage only."),
]


# =========================
# MAG-CXP00002-NP (CXP Serializer IC)
# =========================
CXP_ROWS = [
    # ---- Current consumption (External) ----
    _row("3.3V total supply current (average)", "IVDD3V3_TOTAL_AVG",
         "Sum of analog + digital + PLL supply currents at 3.3 V.",
         EXT, P, "T0007", "IDDD",
         "-", "-", "-", "-", "mA", 5,
         "Sum of analog + digital."),
    _row("3.3V digital supply current (average)", "IVDD3V3_DIG_AVG",
         "Digital + PLL supply current at 3.3 V (single channel).",
         EXT, P, "T0007", "IDDD",
         14.45, 28, 20, 100, "mA", 5,
         "Channel combines DIG+PLL."),
    _row("3.3V analog supply current (average)", "IVDD3V3_ANA_AVG",
         "Analog supply current at 3.3 V, active mode.",
         EXT, P, "T0007", "IDDD",
         "-", "-", 40, 60, "mA", 5,
         "Active mode measurement."),
    _row("3.3V PLL supply current (average)", "IVDD3V3_PLL_AVG",
         "PLL supply current at 3.3 V (included in DIG channel).",
         EXT, C, "T0007", "IDDD",
         "-", "-", "-", "-", "mA", "-",
         "Included in DIG, no separate power supply channel."),
    _row("Total power consumption", "-",
         "Total power consumption derived from IDD x VDD.",
         EXT, C, "-", "-",
         "-", "-", "-", "-", "mW", "-",
         "Derived from IDD x VDD; datasheet typical 271.55 mW."),

    # ---- Logic level inputs (GPIO) ----
    _row("VIL (GPIO input low threshold)", "VIL",
         "GPIO input low threshold voltage.",
         EXT, P, "T0004", "VIL/VIH",
         390, 635, 390, 635, "mV", 5, ""),
    _row("VIH (GPIO input high threshold)", "VIH",
         "GPIO input high threshold voltage.",
         EXT, P, "T0004", "VIL/VIH",
         443, 715, 443, 715, "mV", 5, ""),
    _row("Hysteresis (GPIO)", "Hysteresis",
         "Hysteresis between GPIO VIH and VIL.",
         EXT, P, "T0004", "VIL/VIH",
         38, 117, 38, 117, "mV", 5, ""),

    # ---- Logic level inputs (Video pins) ----
    _row("VIL (Video input low threshold)", "VIL",
         "Video input low threshold voltage.",
         EXT, P, "T0004", "VIL/VIH",
         1.147, 1.689, 1.147, 1.689, "V", 5, ""),
    _row("VIH (Video input high threshold)", "VIH",
         "Video input high threshold voltage.",
         EXT, P, "T0004", "VIL/VIH",
         1.482, 2.282, 1.482, 2.282, "V", 5, ""),
    _row("Hysteresis (Video)", "Hysteresis",
         "Hysteresis between video VIH and VIL.",
         EXT, P, "T0004", "VIL/VIH",
         309, 593, 38, 593, "mV", 5, ""),

    # ---- Data rates ----
    _row("Parallel video data rate", "DRin,par",
         "Maximum parallel video-input data rate accepted by the device (functional test).",
         EXT, P, "T0023", "Camera Test",
         "-", 100, "pass/fail", "-", "MHz", 6,
         "Functional test."),

    # ---- Crystal oscillator ----
    _row("Crystal frequency", "fXTAL",
         "External crystal oscillator frequency range.",
         EXT, P, "T0013", "Oscillator",
         10, 40, 19.5, 20.5, "MHz", 5, ""),

    # ---- CAM LDO outputs ----
    _row("VDD1V8_CAM supply voltage", "V_VDD1V8_CAM",
         "On-chip LDO output for 1.8 V camera supply.",
         EXT, P, "T0012", "Regulators",
         1.7, 1.9, 1.2, 2.4, "V", 5, ""),
    _row("VDD1V2_CAM reference voltage", "V_VDD1V2_CAM",
         "On-chip LDO output for 1.2 V camera reference.",
         EXT, P, "T0011", "Regulators",
         1.15, 1.25, 1.15, 1.25, "V", 5, ""),

    # ---- Internal parameters ----
    _row("Pin continuity (all pins)", "-",
         "ESD-diode voltage drop check on every pin with supplies off.",
         INT, P, "T0001", "Continuity",
         "-", "-", "-", "-", "V", 4, ""),
    _row("Power short (per power rail)", "-",
         "Low-voltage short-circuit check on each supply domain.",
         INT, P, "T0002", "Power Short",
         "-", "-", "-", "-", "mA", 4, ""),
    _row("Pin leakage", "-",
         "Leakage current on input / tristate output pins.",
         INT, P, "T0003", "Pin leakage",
         "-", "-", "-", "-", "nA", 5, ""),
    _row("VOL/VOH (digital outputs)", "-",
         "Digital output low / high voltage under 1 mA load.",
         INT, P, "T0005", "VOH/VOL",
         "-", "-", "-", "-", "V", 5, ""),
    _row("IOL/IOH (digital outputs)", "-",
         "Digital output sink / source current capability.",
         INT, P, "T0006", "IOL/IOH",
         "-", "-", "-", "-", "mA", 5, ""),
    _row("Test buffer offset", "-",
         "Offset of both test-mux buffers measured via GPIO loopback.",
         INT, P, "T0008", "Test Buffer Offset",
         "-", "-", "-", "-", "mV", 5, ""),
    _row("Static signal mux", "-",
         "All static signals available on the test mux stored in the database.",
         INT, P, "T0010", "Static Signal Test",
         "-", "-", "-", "-", "V", 5, ""),
    _row("PLL open-loop v_tune range", "-",
         "PLL open-loop v_tune voltage range.",
         INT, P, "T0014", "PLL Open Loop",
         "-", "-", "-", "-", "V", 5, ""),
    _row("PLL lock (closed-loop)", "-",
         "Closed-loop PLL lock behaviour: lock time, lock indication, jitter flags.",
         INT, P, "T0015", "PLL Lock",
         "-", "-", "-", "-", "-", 6, ""),
    _row("Debug SPI slave (IO SPI Slave)", "-",
         "Validates debug SPI slave interface by register read/write.",
         INT, P, "T0020", "IO SPI Slave",
         "-", "-", "-", "-", "-", 6, ""),
    _row("Camera SPI master (IO SPI Master CAM)", "-",
         "Validates camera SPI master by read/write to FPGA slave.",
         INT, P, "T0020", "IO SPI Master CAM",
         "-", "-", "-", "-", "-", 6, ""),
    _row("GPIO SPI master (IO SPI Master GPIO)", "-",
         "Validates GPIO SPI master by read/write to FPGA slave.",
         INT, P, "T0020", "IO SPI Master GPIO",
         "-", "-", "-", "-", "-", 6, ""),
    _row("CXP uplink receiver sensitivity", "-",
         "Binary search for CXP uplink receiver sensitivity threshold using a function "
         "generator to apply CXP sequences at varying amplitudes.",
         INT, P, "T0022", "CXP Uplink Receiver Sensitivity",
         "-", "-", "-", "-", "mV", 6, ""),
    _row("Internal test patterns", "-",
         "Internal TPG verified across FHD/HD/VGA, 8/10/12-bit, 13 patterns, with frame grabber.",
         INT, P, "T0019", "Internal Test Patterns",
         "-", "-", "-", "-", "-", 6, ""),
    _row("Camera test (external TPG)", "-",
         "External FPGA pattern generator -> video-in -> CXP-out pixel comparison.",
         INT, P, "T0023", "Camera Test",
         "-", "-", "-", "-", "-", 6, ""),
    _row("OTP data retention", "-",
         "Reads OTP and verifies no bits flipped.",
         INT, P, "T0025", "OTP Data Retention",
         "-", "-", "-", "-", "-", 6, ""),
    _row("OTP write + verify", "-",
         "Programs final trim configuration into OTP and verifies successful write.",
         INT, P, "T0026", "OTP Write",
         "-", "-", "-", "-", "-", 6,
         "Performed in TRIM stage only."),
    _row("Block trimming values", "-",
         "Best-fit trim codes for LDOs, voltage / current references.",
         INT, P, "T0021", "Block Trimming",
         "-", "-", "-", "-", "code", 5,
         "Performed in TRIM stage only."),
]


# =========================
# MAG-PSU00001-NP (DC/DC Converter IC)
# =========================
PSU_ROWS = [
    # ---- Power ----
    _row("Input voltage supply range", "PVin, Vin",
         "Operating input voltage range of the converter (test condition).",
         EXT, D, "-", "-",
         5, 11, "-", "-", "V", "-",
         "Test condition."),
    _row("Output current (PCB in air)", "Iout",
         "Maximum output current under free-air PCB conditions.",
         EXT, D, "-", "-",
         "-", 1, "-", "-", "A", "-",
         "Test condition."),
    _row("Output power (PCB in air)", "Pout",
         "Maximum output power under free-air PCB conditions.",
         EXT, D, "-", "-",
         "-", 2, "-", "-", "W", "-",
         "Test condition."),

    # ---- PWM ----
    _row("Maximum duty cycle", "DMax",
         "Maximum PWM duty cycle achievable by the controller.",
         EXT, S, "-", "-",
         "-", "-", "-", "-", "%", "-",
         "Cannot be measured, sensitive node; verified by simulation."),
    _row("Minimum duty cycle", "DMin",
         "Minimum PWM duty cycle achievable by the controller.",
         EXT, S, "-", "-",
         "-", "-", "-", "-", "%", "-",
         "Cannot be measured, sensitive node; verified by simulation."),

    # ---- Error Amplifier ----
    _row("DC Gain", "DCG",
         "Error amplifier DC gain.",
         EXT, S, "-", "-",
         "-", "-", "-", "-", "dB", "-",
         "Cannot be measured, sensitive node."),
    _row("Unity Gain-Bandwidth", "UGBW",
         "Error amplifier unity gain bandwidth.",
         EXT, S, "-", "-",
         "-", "-", "-", "-", "MHz", "-",
         "Cannot be measured, sensitive node."),
    _row("Slew Rate", "SR",
         "Error amplifier slew rate.",
         EXT, S, "-", "-",
         "-", "-", "-", "-", "V/us", "-",
         "Cannot be measured, sensitive node."),

    # ---- UVLO ----
    _row("Vin start threshold", "VinStartTh",
         "Input voltage at which the converter enters operation.",
         EXT, P, "T0009", "Under Voltage Lock Out",
         "-", "-", 4.4, 4.8, "V", 5, ""),
    _row("Vin stop threshold", "VinStopTh",
         "Input voltage at which the converter exits operation (UVLO).",
         EXT, P, "T0009", "Under Voltage Lock Out",
         "-", "-", 4.12, 4.53, "V", 5, ""),

    # ---- Enable ----
    _row("Enable start threshold", "EnStartTh",
         "Enable pin rising threshold to start the converter.",
         EXT, P, "T0006", "Input threshold",
         "-", "-", 770, 900, "mV", 5, ""),
    _row("Enable stop threshold", "EnStopTh",
         "Enable pin falling threshold to stop the converter.",
         EXT, P, "T0006", "Input threshold",
         "-", "-", 670, 780, "mV", 5, ""),
    _row("Enable pin series resistance", "EnSerRes",
         "Internal series resistance on the enable pin.",
         EXT, D, "-", "-",
         "-", "-", "-", "-", "kOhm", "-",
         "Guaranteed by design."),

    # ---- Protections ----
    _row("Over-current protection peak level", "OCPpk",
         "Peak inductor current at which OCP trips.",
         EXT, D, "-", "-",
         "-", "-", "-", "-", "A", "-",
         "Cannot be triggered with test setup; OCP functionality not covered within test plan."),
    _row("Over-current protection average level", "OCPavg",
         "Average inductor current at which OCP trips.",
         EXT, D, "-", "-",
         "-", "-", "-", "-", "A", "-",
         "Cannot be triggered with test setup; OCP functionality not covered within test plan."),
    _row("Over-temperature start threshold", "OTPStartTh",
         "Junction temperature at which over-temperature protection trips.",
         EXT, C, "-", "-",
         "-", "-", "-", "-", "degC", "-",
         "Only tested during characterization."),
    _row("Over-temperature stop threshold", "OTPStopTh",
         "Junction temperature at which over-temperature protection releases.",
         EXT, C, "-", "-",
         "-", "-", "-", "-", "degC", "-",
         "Only tested during characterization."),

    # ---- Soft Start ----
    _row("Soft-start duration", "SSt",
         "Time for output to ramp from 0 V to regulation target.",
         EXT, P, "T0012", "Startup behaviour",
         "-", 440, "-", 1500, "us", 5, ""),

    # ---- Power Good ----
    _row("Output over-voltage PGood threshold", "OV",
         "Output over-voltage threshold at which the PGood flag is de-asserted (% of target).",
         EXT, P, "T0010", "Output voltage monitor",
         "-", "-", 2.5, 10.5, "%", 5, ""),
    _row("Output under-voltage PGood threshold", "UV",
         "Output under-voltage threshold at which the PGood flag is de-asserted (% of target).",
         EXT, P, "T0010", "Output voltage monitor",
         "-", "-", -10.5, -2.5, "%", 5, ""),

    # ---- PTAT ----
    _row("PTAT temperature slope", "PTAT",
         "PTAT voltage slope vs die temperature.",
         EXT, C, "-", "-",
         "-", "-", "-", "-", "mV/degC", "-",
         "Only tested during characterization."),

    # ---- Internal ----
    _row("Power short (power terminals)", "-",
         "Shorts-to-ground check on the converter power terminals.",
         INT, P, "T0002", "Power Short",
         "-", "-", "-", "-", "A", 4,
         "CONN stage."),
    _row("Load short (load terminals)", "-",
         "Shorts-to-ground / power check on the converter load terminals.",
         INT, P, "T0002", "Load Short",
         "-", "-", "-", "-", "A", 4,
         "CONN stage."),
    _row("Input pin leakage", "-",
         "Leakage current on accessible input pins.",
         INT, P, "T0003", "Leakage",
         "-", "-", "-", "-", "nA", 5, ""),
    _row("Power-good output drive strength", "-",
         "Sink current capability of the PGood output.",
         INT, P, "T0006", "Output drive-strength",
         "-", "-", "-", "-", "mA", 5, ""),
    _row("On-chip LDO regulator voltage", "-",
         "Internal LDO output voltage and load regulation.",
         INT, P, "T0011", "On-chip regulator",
         "-", "-", "-", "-", "V", 5, ""),
    _row("Over-temperature protection flag (FT)", "-",
         "OTP flag via PGood and PTAT voltage monitoring during FT.",
         INT, P, "T0015", "Over Temperature Protection",
         "-", "-", "-", "-", "mV", 5, ""),
    _row("Inv_Enable functionality", "-",
         "Verifies the enable-pin polarity can be inverted via Inv_Enable configuration.",
         INT, P, "T0018", "Inv_Enable functionality",
         "-", "-", "-", "-", "-", 6, ""),
    _row("Line regulation", "-",
         "Output voltage variation vs input voltage (Vout: 0.9/1.2/1.5/1.8/2.0/2.5/3.3 V, Vmin/Vmax).",
         INT, P, "T0027", "Line regulation",
         "-", "-", "-", "-", "mV/V", 5, ""),
    _row("Load regulation", "-",
         "Output voltage variation vs load current (Iload: 0.1-1.1 A across 7 Vout settings).",
         INT, P, "T0028", "Load regulation",
         "-", "-", "-", "-", "mV/A", 5, ""),
    _row("Efficiency", "-",
         "Converter efficiency across loads/voltages, half-switched mode ON/OFF.",
         INT, P, "T0029", "Efficiency",
         "-", "-", "-", "-", "%", 5, ""),
]


# =========================
# MAG-VIS100xx-N (Electronic Vision Module)
# =========================
VIS_ROWS = [
    # ---- Datasheet parameters ----
    _row("Module supply current", "I_SUPPLY_9V",
         "Total module supply current drawn from the 9 V input at RT / Vnom.",
         EXT, P, "T0100", "Power consumption",
         160, 257, 160, 257, "mA", 5, ""),
    _row("AC output impedance (HD-BNC)", "Z_out_AC",
         "AC output impedance at the HD-BNC CXP connector.",
         EXT, D, "-", "-",
         "-", "-", "-", "-", "Ohm", "-",
         "Datasheet typical: 75 Ohm. Guaranteed by PCB/connector design."),

    # ---- Internal / functional ----
    _row("Visual inspection", "-",
         "Manual pre-test inspection for cracks, solder defects, contamination on the integrated stack.",
         INT, P, "T0101", "Visual inspection",
         "-", "-", "-", "-", "-", 3, ""),
    _row("CXP core functionality + video data", "-",
         "Lock-bit check, register read/write over CoaXPress, internal TPG verified via frame grabber.",
         INT, P, "T0102", "CXP Core functionality and video data check",
         "-", "-", "-", "-", "-", 6, ""),
    _row("IMG core functionality + video data", "-",
         "PLL status check, y-axis flip image-manipulation logic, internal TPG across video modes / frame rates.",
         INT, P, "T0103", "IMG Core functionality and video data check",
         "-", "-", "-", "-", "-", 6, ""),
    _row("IMG trigger modes", "-",
         "CXP hardware trigger flow, free-run burst (30 frames) and software trigger burst modes.",
         INT, P, "T0104", "IMG Trigger mode",
         "-", "-", "-", "-", "-", 6, ""),
    _row("MAG-DRV output capabilities (MAG-VIS1001x only)", "-",
         "Steady-state (ON/OFF) and PWM modes on MAG-DRV outputs, verified with scope and mux card.",
         INT, P, "T0105", "MAG-DRV Capabilities",
         "-", "-", "-", "-", "-", 6,
         "Variant-specific test (MAG-VIS1001x)."),
    _row("MAG-DRV motor drive (MAG-VIS1002x only)", "-",
         "Iris / Zoom / Focus motor drive directions and safe disable via drv_enable register.",
         INT, P, "T0106", "MAG-DRV Motor drive capabilities",
         "-", "-", "-", "-", "-", 6,
         "Variant-specific test (MAG-VIS1002x)."),
]


# =========================
# Workbook assembly
# =========================
PRODUCT_SHEETS = [
    ("MAG-IMG002x1-NC", "MAG-IMG002x1-NC (Full HD CMOS Image Sensor)", IMG_ROWS),
    ("MAG-CXP00002-NP", "MAG-CXP00002-NP (CoaXPress Serializer ASIC)",  CXP_ROWS),
    ("MAG-PSU00001-NP", "MAG-PSU00001-NP (10 W Synchronous Step-Down DC/DC Converter)", PSU_ROWS),
    ("MAG-VIS100xx-N",  "MAG-VIS100xx-N (Electronic Vision Module)",   VIS_ROWS),
]

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="305496")
SECTION_FILL = PatternFill("solid", fgColor="D9E1F2")
EXT_FILL = PatternFill("solid", fgColor="E2EFDA")
INT_FILL = PatternFill("solid", fgColor="FFF2CC")
WRAP = Alignment(vertical="top", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

COL_WIDTHS = {
    "Parameter": 40,
    "Datasheet Label": 18,
    "Description": 55,
    "Class": 10,
    "Test Level": 11,
    "Test ID": 9,
    "Test Procedure": 24,
    "LSL": 10,
    "USL": 10,
    "LTL": 10,
    "UTL": 10,
    "Unit": 10,
    "Bin (fail)": 11,
    "Comment": 50,
}


def build_overview_sheet(wb):
    ws = wb.create_sheet("Overview", 0)
    ws["A1"] = "Vision Series - Compliance Matrix"
    ws["A1"].font = Font(bold=True, size=16, color="305496")
    ws.merge_cells("A1:D1")

    rows = [
        ("Source document", "Vision_Series_Production_Test_Plan.md (rev 1.3, 2026-04-22)"),
        ("Generated by",    Path(__file__).name),
        ("Products covered",
         "MAG-IMG002x1-NC, MAG-CXP00002-NP, MAG-PSU00001-NP, MAG-VIS100xx-N"),
        ("", ""),
        ("Column", "Meaning"),
        ("Parameter",       "Name of the measured parameter."),
        ("Datasheet Label", "Symbol used in the product datasheet (blank when parameter is internal)."),
        ("Description",     "Short description extracted from the test plan / datasheet text."),
        ("Class",
         "External = parameter is published in the product datasheet; "
         "Internal = parameter is monitored internally only for process / yield control."),
        ("Test Level",
         "D = By design (guaranteed without measurement); "
         "S = Simulation (validated in SPICE / top-level sim); "
         "C = Characterization (measured on a sample, not every unit); "
         "P = Production test (measured on every unit)."),
        ("Test ID",         "T-ID from the parameter traceability tables of the test plan."),
        ("Test Procedure",  "Test procedure name from the test-plan procedure tables."),
        ("LSL / USL",       "Lower / Upper Specification Limit (= datasheet DS Min / DS Max)."),
        ("LTL / UTL",       "Lower / Upper Test Limit applied during production screening (from Chapter 5 tables)."),
        ("Unit",            "Physical unit of the parameter."),
        ("Bin (fail)",
         "Bin number to which a unit is classified when the parameter falls outside its PROD limits. "
         "See the 'Bin List' sheet for full definitions."),
        ("Comment",         "Notes taken from the test plan comments column."),
    ]
    for row_idx, (a, b) in enumerate(rows, start=3):
        ws.cell(row=row_idx, column=1, value=a).font = Font(bold=(row_idx == 7 or b == "Meaning"))
        ws.cell(row=row_idx, column=2, value=b).alignment = Alignment(wrap_text=True, vertical="top")

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 110


def build_bin_sheet(wb):
    ws = wb.create_sheet("Bin List")
    ws.append(["Bin #", "Bin Name", "Result", "Description"])
    for col in range(1, 5):
        c = ws.cell(row=1, column=col)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER
    for b in BIN_LIST:
        ws.append(list(b))

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=4):
        for cell in row:
            cell.alignment = WRAP
            cell.border = BORDER

    widths = {1: 8, 2: 14, 3: 12, 4: 95}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 22


def build_product_sheet(wb, sheet_name, title, rows):
    ws = wb.create_sheet(sheet_name)

    ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=14, color="305496")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))

    for col_idx, header in enumerate(HEADERS, start=1):
        c = ws.cell(row=3, column=col_idx, value=header)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER

    for r_idx, row in enumerate(rows, start=4):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = WRAP if HEADERS[c_idx - 1] in ("Parameter", "Description", "Comment", "Test Procedure") else CENTER
            cell.border = BORDER

        klass = row[3]
        fill = EXT_FILL if klass == EXT else INT_FILL
        ws.cell(row=r_idx, column=4).fill = fill

    for col_idx, header in enumerate(HEADERS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS[header]

    ws.freeze_panes = "A4"
    ws.row_dimensions[3].height = 28

    first_col = "A"
    last_col = get_column_letter(len(HEADERS))
    last_row = 3 + len(rows)
    table_ref = f"{first_col}3:{last_col}{last_row}"
    table = Table(displayName=f"tbl_{sheet_name.replace('-', '_')}", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleLight1",
        showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False,
    )
    ws.add_table(table)


def main() -> int:
    wb = Workbook()
    wb.remove(wb.active)

    build_overview_sheet(wb)
    build_bin_sheet(wb)
    for sheet_name, title, rows in PRODUCT_SHEETS:
        build_product_sheet(wb, sheet_name, title, rows)

    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_FILE)

    print(f"Wrote: {OUTPUT_FILE}")
    print(f"Sheets: {wb.sheetnames}")
    for sheet_name, _, rows in PRODUCT_SHEETS:
        ext = sum(1 for r in rows if r[3] == EXT)
        intn = sum(1 for r in rows if r[3] == INT)
        print(f"  - {sheet_name}: {len(rows)} rows ({ext} External / {intn} Internal)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
